import os
import re
import time
import shutil
import unicodedata
import math
import requests
from pathlib import Path
from datetime import datetime, timedelta, date

import pandas as pd

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.edge.options import Options
from selenium.webdriver.edge.service import Service as EdgeService
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC

from selenium.common.exceptions import (
    TimeoutException,
    StaleElementReferenceException,
    NoSuchElementException,
    WebDriverException,
    ElementClickInterceptedException,
)

from selenium.webdriver.edge import service as edge_service

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


# ============================================================
# CONFIGURAÇÕES
# ============================================================

URL_LOGIN = "https://www.autovision.com.br/v3/"

URL_RELATORIO_POSICAO = (
    "https://www.autovision.com.br/v3/"
    "modulos/relatorios/"
    "relatorio_posicao.php"
)


# ============================================================
# GOOGLE APPS SCRIPT - ENVIO DE RESULTADOS
# ============================================================

GOOGLE_SCRIPT_URL = os.environ.get(
    "GOOGLE_SCRIPT_URL",
    "https://script.google.com/macros/s/AKfycbxREO251djkCbe1HKo8wIxDhXM9CVeaBsMF3lzphYDTjM0272WTzne3PnFoMl9sUNWRhw/exec"
)



# ============================================================
# FONTE OFICIAL DE ABASTECIMENTO
# ============================================================

URL_ABASTECIMENTO = (
    "https://docs.google.com/spreadsheets/d/e/"
    "2PACX-1vTNyx3mdkh9hF027_l61y7O7dwYr_gF5ofFwi0mzRY0eNQuKCu3KR3peiCn7Q_832YRjaxR3rqxQGaB/"
    "pub?gid=1282350705&single=true&output=csv"
)

URL_ABASTECIMENTO_EDIT = (
    "https://docs.google.com/spreadsheets/d/"
    "1NdRe71fCKND18sIOdcivRhhwTJYS4JnMjJfGUaAbZuM/"
    "edit?gid=1282350705#gid=1282350705"
)


# ============================================================
# TELEMETRIA OCIOSIDADE
#
# AS PLACAS SERÃO LIDAS AUTOMATICAMENTE DE:
#
# Aba: Geral
# Coluna: PLACA
# ============================================================

URL_TELEMETRIA_OCIOSIDADE = (
    "https://docs.google.com/spreadsheets/d/e/"
    "2PACX-1vQU22jWG-2LuGLDuLYvVQoBbv5JdArq9WUwGcJ3znHPZWHrFABji3IFNFwYCtVX7u8uo-Rd7YJFb9fZ/"
    "pub?gid=756259345&single=true&output=csv"
)

URL_TELEMETRIA_OCIOSIDADE_EDIT = (
    "https://docs.google.com/spreadsheets/d/"
    "1q-SxCJ4C97uEzuPUyvykM8maMw0VxF540nwfK7bv0SA/"
    "edit?gid=756259345#gid=756259345"
)

ABA_TELEMETRIA_OCIOSIDADE = "Geral"

COLUNA_PLACA_TELEMETRIA = "PLACA"


# ============================================================
# PLACAS
#
# NÃO PREENCHER MANUALMENTE.
#
# Será preenchida automaticamente pela função:
#
# carregar_placas_telemetria()
# ============================================================

PLACAS = []


# ============================================================
# PASTAS
# ============================================================

PASTA_BASE = Path.cwd()

PASTA_HTML = (
    PASTA_BASE /
    "relatorios_posicao"
)

PASTA_EXCEL_NEXUS = (
    PASTA_BASE /
    "analises_nexus"
)

ARQUIVO_UNIFICADO = (
    PASTA_BASE /
    "unificado_desvios.xlsx"
)

PASTA_HTML.mkdir(
    parents=True,
    exist_ok=True
)

PASTA_EXCEL_NEXUS.mkdir(
    parents=True,
    exist_ok=True
)


# ============================================================
# MICROSOFT EDGE DRIVER
#
# O msedgedriver será baixado automaticamente pelo Selenium.
# ============================================================

PASTA_DRIVER = (
    PASTA_BASE /
    "drivers"
)

PASTA_DRIVER.mkdir(
    parents=True,
    exist_ok=True
)


# ============================================================
# AUDITORIA LOCAL DOS RELATÓRIOS AUTOVISION
# O Nexus não é mais utilizado para identificar os desvios.
# ============================================================

PADRAO_ARQUIVO = re.compile(
    r"^(?P<placa>[A-Z0-9-]+)_(?P<data>\d{8})_(?P<hora>\d{6})$",
    re.IGNORECASE,
)

PASTA_SAIDA_AUDITORIA = (
    PASTA_BASE /
    "analises_telemetria"
)

ARQUIVO_AUDITORIA = (
    PASTA_SAIDA_AUDITORIA /
    "auditoria_telemetria.xlsx"
)

PASTA_SAIDA_AUDITORIA.mkdir(
    parents=True,
    exist_ok=True
)

TEMPO_MINIMO_PARADA = 15

GAP_MAX_MINUTOS = 20

RAIO_PARADA_METROS = 100

RAIO_BUSCA_POI_METROS = 300

RAIO_DIVERGENCIA_ENDERECO_METROS = 200

TOLERANCIA_ABASTECIMENTO_MIN = 3

TOLERANCIA_IGNICAO_ABASTECIMENTO_MIN = 3










# ============================================================
# NOME SEGURO PARA ARQUIVOS
# ============================================================

def nome_seguro(valor):
    """
    Remove caracteres inválidos para nomes de arquivos
    no Windows/Linux e mantém somente caracteres seguros.
    """

    texto = str(valor or "").strip()

    if not texto:
        return "arquivo"

    # Remove caracteres inválidos
    texto = re.sub(
        r'[<>:"/\\|?*]',
        "_",
        texto
    )

    # Remove caracteres de controle
    texto = re.sub(
        r"[\x00-\x1f\x7f]",
        "_",
        texto
    )

    # Evita espaços repetidos
    texto = re.sub(
        r"\s+",
        "_",
        texto
    )

    # Remove pontos/espaços no final
    texto = texto.rstrip(". ")

    return texto or "arquivo"


















# ============================================================
# GOOGLE PLACES API - OPCIONAL / ÚLTIMO RECURSO
# ============================================================

GOOGLE_PLACES_API_KEY = os.environ.get(
    "GOOGLE_PLACES_API_KEY",
    "AIzaSyBtJmgG71Efuw_JEN8_-nQHps53K9c-RBY"
)

OVERPASS_URL = os.environ.get(
    "OVERPASS_URL",
    "https://overpass.kumi.systems/api/interpreter"
)

OVERPASS_URLS = [
    OVERPASS_URL,
    "https://overpass-api.de/api/interpreter",
    "https://overpass.private.coffee/api/interpreter",
]

NOMINATIM_USER_AGENT = os.environ.get(
    "NOMINATIM_USER_AGENT",
    "COMPESA-AuditoriaFrota/1.0"
)

OSM_MIN_INTERVAL_S = 1.1

_ULTIMA_CONSULTA_OSM = 0.0

_CACHE_LOCAIS = {}


CRITICIDADE_MAP = {
    "lodging": "A",
    "hotel": "A",
    "motel": "A",
    "beach": "A",
    "bar": "A",
    "night_club": "A",
    "shopping_mall": "A",
    "campground": "A",
    "rv_park": "A",

    "restaurant": "B",
    "meal_takeaway": "B",
    "supermarket": "B",
    "gym": "B",
    "school": "B",
    "university": "B",
    "movie_theater": "B",
    "performing_arts_theater": "B",
    "hospital": "B",
    "doctor": "B",
    "place_of_worship": "B",
    "airport": "B",
    "bus_station": "B",

    "park": "C",
    "bank": "C",
    "atm": "C",
    "bakery": "C",
    "grocery_or_supermarket": "C",
    "cafe": "C",
    "ice_cream_shop": "C",
    "gas_station": "C",
    "parking": "C",
    "car_repair": "C",
    "car_wash": "C",
}


OSM_CRITICIDADE_MAP = {
    ("tourism", "hotel"): "A",
    ("tourism", "motel"): "A",
    ("tourism", "guest_house"): "A",
    ("tourism", "hostel"): "A",
    ("tourism", "apartment"): "A",
    ("tourism", "chalet"): "A",
    ("tourism", "camp_site"): "A",
    ("tourism", "caravan_site"): "A",
    ("tourism", "resort"): "A",

    ("natural", "beach"): "A",

    ("amenity", "bar"): "A",
    ("amenity", "nightclub"): "A",
    ("amenity", "pub"): "A",
    ("amenity", "casino"): "A",
    ("amenity", "love_hotel"): "A",

    ("shop", "mall"): "A",
    ("shop", "department_store"): "A",

    ("amenity", "restaurant"): "B",
    ("amenity", "fast_food"): "B",

    ("shop", "supermarket"): "B",

    ("leisure", "fitness_centre"): "B",

    ("amenity", "school"): "B",
    ("amenity", "university"): "B",
    ("amenity", "college"): "B",

    ("amenity", "cinema"): "B",
    ("amenity", "theatre"): "B",

    ("amenity", "hospital"): "B",
    ("amenity", "clinic"): "B",
    ("amenity", "doctors"): "B",

    ("amenity", "place_of_worship"): "B",

    ("aeroway", "aerodrome"): "B",

    ("amenity", "bus_station"): "B",

    ("leisure", "park"): "C",
    ("amenity", "bank"): "C",
    ("amenity", "atm"): "C",

    ("shop", "bakery"): "C",
    ("shop", "convenience"): "C",
    ("shop", "grocery"): "C",

    ("amenity", "cafe"): "C",
    ("amenity", "ice_cream"): "C",

    ("shop", "ice_cream"): "C",

    ("amenity", "fuel"): "C",
    ("amenity", "parking"): "C",

    ("shop", "car_repair"): "C",
    ("amenity", "car_wash"): "C",
}


OSM_TAG_KEYS = [
    "amenity",
    "shop",
    "tourism",
    "leisure",
    "aeroway"
]

CRITICIDADE_PESO = {
    "A": 3,
    "B": 2,
    "C": 1
}

CRITICIDADE_COR = {
    "A": "🔴",
    "B": "🟡",
    "C": "🔵"
}


# ============================================================
# TEMPOS
# ============================================================

TIMEOUT = 30

TIMEOUT_RELATORIO = 120

TIMEOUT_DOWNLOAD = 90

INTERVALO_ENTRE_PLACAS = 1.0

INTERVALO_ENTRE_ANALISES = 2.0


# ============================================================
# AUTOVISION
# ============================================================

USUARIO = os.environ.get(
    "AUTOVISION_USUARIO",
    "nayarasilva"
)

SENHA = os.environ.get(
    "AUTOVISION_SENHA",
    "20245"
)


# ============================================================
# NEXUS
# ============================================================

NEXUS_EMAIL = os.environ.get(
    "NEXUS_EMAIL",
    "nayarasilva@compesa.com.br"
)

NEXUS_SENHA = os.environ.get(
    "NEXUS_SENHA",
    "acessologin"
)


# ============================================================
# NORMALIZAR TEXTO
# ============================================================

def normalizar_texto(valor):

    if valor is None:
        return ""

    valor = str(valor)

    valor = unicodedata.normalize(
        "NFKD",
        valor
    )

    valor = "".join(
        c
        for c in valor
        if not unicodedata.combining(c)
    )

    valor = re.sub(
        r"\s+",
        " ",
        valor
    )

    return valor.strip().upper()


# ============================================================
# NORMALIZAR PLACA
# ============================================================

def normalizar_placa(valor):

    if valor is None:
        return ""

    valor = normalizar_texto(
        valor
    )

    valor = re.sub(
        r"[^A-Z0-9]",
        "",
        valor
    )

    return valor


# ============================================================
# CARREGAR PLACAS DA TELEMETRIA OCIOSIDADE
#
# Fonte:
# Google Sheets -> aba Geral -> coluna PLACA
#
# Não utiliza mais:
#
# PLACAS = [...]
# ============================================================

def carregar_placas_telemetria():

    global PLACAS

    print()
    print("=" * 70)
    print("CARREGANDO PLACAS DA TELEMETRIA OCIOSIDADE")
    print("=" * 70)

    print(
        "Fonte:",
        URL_TELEMETRIA_OCIOSIDADE
    )

    try:

        # ----------------------------------------------------
        # A URL publicada já aponta para a planilha/aba Geral.
        # ----------------------------------------------------

        df = pd.read_csv(
            URL_TELEMETRIA_OCIOSIDADE,
            dtype=str
        )

    except Exception as erro:

        raise RuntimeError(
            "Não foi possível carregar a planilha "
            "Telemetria Ociosidade.\n"
            f"URL: {URL_TELEMETRIA_OCIOSIDADE}\n"
            f"Erro: {erro}"
        )

    if df.empty:

        raise RuntimeError(
            "A planilha Telemetria Ociosidade "
            "foi carregada, porém está vazia."
        )

    # --------------------------------------------------------
    # NORMALIZAR NOMES DAS COLUNAS
    # --------------------------------------------------------

    mapa_colunas = {}

    for coluna in df.columns:

        coluna_normalizada = normalizar_texto(
            coluna
        )

        coluna_normalizada = re.sub(
            r"[^A-Z0-9]",
            "",
            coluna_normalizada
        )

        mapa_colunas[
            coluna_normalizada
        ] = coluna

    chave_placa = re.sub(
        r"[^A-Z0-9]",
        "",
        normalizar_texto(
            COLUNA_PLACA_TELEMETRIA
        )
    )

    coluna_placa = mapa_colunas.get(
        chave_placa
    )

    if coluna_placa is None:

        raise RuntimeError(
            "A coluna 'PLACA' não foi encontrada "
            "na aba Geral da Telemetria Ociosidade.\n"
            f"Colunas encontradas: {list(df.columns)}"
        )

    # --------------------------------------------------------
    # EXTRAIR PLACAS
    # --------------------------------------------------------

    placas = []

    for valor in df[coluna_placa]:

        placa = normalizar_placa(
            valor
        )

        if placa:

            placas.append(
                placa
            )

    # --------------------------------------------------------
    # REMOVER DUPLICADAS PRESERVANDO ORDEM
    # --------------------------------------------------------

    PLACAS = list(
        dict.fromkeys(
            placas
        )
    )

    if not PLACAS:

        raise RuntimeError(
            "Nenhuma placa válida foi encontrada "
            "na coluna PLACA da aba Geral."
        )

    # --------------------------------------------------------
    # EXIBIR RESULTADO
    # --------------------------------------------------------

    print()
    print("✓ Planilha carregada.")
    print(
        f"✓ Coluna utilizada: {coluna_placa}"
    )

    print(
        f"✓ Total de placas encontradas: "
        f"{len(PLACAS)}"
    )

    print()
    print("PLACAS CARREGADAS:")

    for indice, placa in enumerate(
        PLACAS,
        start=1
    ):

        print(
            f"{indice}. {placa}"
        )

    print("=" * 70)

    return PLACAS


# ============================================================
# VALIDAR PLACAS
#
# Agora a validação é feita depois da leitura da planilha.
# ============================================================


def carregar_placas_telemetria_ociosidade():
    print()
    print("=" * 70)
    print("CARREGANDO PLACAS - TELEMETRIA OCIOSIDADE")
    print("=" * 70)

    try:
        resposta = requests.get(
            URL_TELEMETRIA_OCIOSIDADE,
            timeout=60,
            headers={
                "User-Agent": "Mozilla/5.0 COMPESA-AuditoriaFrota/1.0"
            },
        )

        resposta.raise_for_status()

        conteudo = resposta.content

        try:
            texto = conteudo.decode("utf-8-sig")
        except UnicodeDecodeError:
            texto = conteudo.decode(
                "latin-1",
                errors="replace"
            )

        from io import StringIO

        try:
            df = pd.read_csv(
                StringIO(texto),
                sep=None,
                engine="python",
                dtype=object,
            )
        except Exception:
            df = pd.read_csv(
                StringIO(texto),
                sep=",",
                dtype=object,
            )

        if df.empty:
            raise RuntimeError(
                "A Telemetria Ociosidade retornou uma planilha vazia."
            )

        df.columns = [
            str(coluna).strip()
            for coluna in df.columns
        ]

        coluna_placa = encontrar_coluna(
            df,
            [
                "PLACA",
                "Placa",
                "PLACA VEICULO",
                "PLACA VEÍCULO",
            ],
        )

        if coluna_placa is None:
            raise RuntimeError(
                "A coluna 'PLACA' não foi encontrada "
                "na Telemetria Ociosidade."
            )

        placas = []

        for valor in df[coluna_placa]:

            placa = normalizar_placa(valor)

            if placa and placa not in placas:
                placas.append(placa)

        if not placas:
            raise RuntimeError(
                "Nenhuma placa válida foi encontrada "
                "na coluna PLACA da Telemetria Ociosidade."
            )

        print(
            f"✓ Coluna encontrada: {coluna_placa}"
        )

        print(
            f"✓ Total de placas encontradas: {len(placas)}"
        )

        print(
            "✓ Placas:"
        )

        for placa in placas:
            print(
                f"   - {placa}"
            )

        return placas

    except Exception as erro:

        print()
        print(
            "❌ ERRO AO CARREGAR PLACAS "
            "DA TELEMETRIA OCIOSIDADE"
        )

        print(
            f"   {erro}"
        )

        raise


# ============================================================
# CRIAR DRIVER EDGE
#
# Compatível com:
# - Windows local
# - GitHub Actions / Linux
#
# O Selenium Manager gerencia automaticamente o driver.
# ============================================================

def criar_driver():

    print()
    print("=" * 70)
    print("INICIANDO MICROSOFT EDGE")
    print("=" * 70)

    options = Options()

    # ========================================================
    # AMBIENTE GITHUB ACTIONS / LINUX
    # ========================================================

    # GitHub Actions normalmente não possui interface gráfica.
    # O Edge precisa funcionar em modo headless.
    options.add_argument("--headless=new")

    # Evita problemas de sandbox no ambiente Linux do runner.
    options.add_argument("--no-sandbox")

    # Evita problemas de memória compartilhada (/dev/shm).
    options.add_argument("--disable-dev-shm-usage")

    # Evita problemas relacionados à GPU.
    options.add_argument("--disable-gpu")

    # Tamanho da janela virtual do navegador.
    options.add_argument("--window-size=1920,1080")

    # ========================================================
    # ESTABILIDADE
    # ========================================================

    options.add_argument("--disable-notifications")

    options.add_argument("--disable-popup-blocking")

    options.add_argument(
        "--disable-blink-features=AutomationControlled"
    )

    options.add_argument(
        "--disable-background-networking"
    )

    options.add_argument(
        "--disable-background-timer-throttling"
    )

    options.add_argument(
        "--disable-backgrounding-occluded-windows"
    )

    options.add_argument(
        "--disable-renderer-backgrounding"
    )

    options.add_argument(
        "--no-first-run"
    )

    options.add_argument(
        "--no-default-browser-check"
    )

    # ========================================================
    # REDUZ INDÍCIOS DE AUTOMAÇÃO
    # ========================================================

    options.add_experimental_option(
        "excludeSwitches",
        [
            "enable-automation"
        ]
    )

    # ========================================================
    # DOWNLOAD DOS HTMLs
    # ========================================================

    PASTA_HTML.mkdir(
        parents=True,
        exist_ok=True
    )

    preferencias = {
        "download.default_directory":
            str(PASTA_HTML.resolve()),

        "download.prompt_for_download":
            False,

        "download.directory_upgrade":
            True,

        "safebrowsing.enabled":
            True,

        "profile.default_content_setting_values.automatic_downloads":
            1,

        "download_restrictions":
            0,
    }

    options.add_experimental_option(
        "prefs",
        preferencias
    )

    # ========================================================
    # INICIAR EDGE
    # ========================================================

    try:

        print("Iniciando Edge via Selenium Manager...")

        driver = webdriver.Edge(
            options=options
        )

    except Exception as erro:

        print()
        print("=" * 70)
        print("❌ ERRO AO INICIAR MICROSOFT EDGE")
        print("=" * 70)
        print(erro)

        raise RuntimeError(
            "Não foi possível iniciar o Microsoft Edge.\n\n"
            "Verifique se o workflow do GitHub Actions "
            "instalou o Microsoft Edge corretamente.\n\n"
            f"Erro original: {erro}"
        ) from erro

    # ========================================================
    # CONFIGURAÇÕES DO DRIVER
    # ========================================================

    try:

        driver.set_page_load_timeout(
            120
        )

    except Exception:
        pass

    try:

        driver.set_script_timeout(
            120
        )

    except Exception:
        pass

    # ========================================================
    # LIBERAR DOWNLOADS EM HEADLESS
    # ========================================================

    try:

        driver.execute_cdp_cmd(
            "Browser.setDownloadBehavior",
            {
                "behavior": "allow",
                "downloadPath": str(
                    PASTA_HTML.resolve()
                ),
            }
        )

        print(
            "✓ Download automático habilitado."
        )

    except Exception as erro:

        print(
            f"⚠ Não foi possível configurar "
            f"download via CDP: {erro}"
        )

    # ========================================================
    # RESULTADO
    # ========================================================

    print(
        "✓ Microsoft Edge iniciado."
    )

    print(
        "✓ Modo headless ativo."
    )

    print(
        "✓ Ambiente compatível com GitHub Actions."
    )

    print(
        "✓ Selenium Manager responsável pelo WebDriver."
    )

    print(
        f"✓ Pasta HTML: {PASTA_HTML.resolve()}"
    )

    return driver


# ============================================================
# OBTER PERÍODO
# ============================================================

def obter_periodo():

    hoje = datetime.now()

    print()
    print("=" * 70)
    print("CALCULANDO PERÍODO")
    print("=" * 70)

    print(
        "Data atual:",
        hoje.strftime(
            "%d/%m/%Y %H:%M:%S"
        )
    )

    if hoje.weekday() == 0:

        inicio = (
            hoje -
            timedelta(days=3)
        ).replace(
            hour=0,
            minute=0,
            second=0,
            microsecond=0
        )

        fim = (
            hoje -
            timedelta(days=1)
        ).replace(
            hour=23,
            minute=59,
            second=59,
            microsecond=0
        )

        print(
            "Hoje é segunda-feira."
        )

        print(
            "Período: sexta-feira até domingo."
        )

    else:

        ontem = (
            hoje -
            timedelta(days=1)
        )

        inicio = ontem.replace(
            hour=0,
            minute=0,
            second=0,
            microsecond=0
        )

        fim = ontem.replace(
            hour=23,
            minute=59,
            second=59,
            microsecond=0
        )

        print(
            "Período: dia anterior."
        )

    print(
        "INÍCIO:",
        inicio.strftime(
            "%d/%m/%Y %H:%M:%S"
        )
    )

    print(
        "FIM:",
        fim.strftime(
            "%d/%m/%Y %H:%M:%S"
        )
    )

    return inicio, fim


# ============================================================
# LOGIN AUTOVISION
# ============================================================

def fazer_login(driver):

    print()
    print("=" * 70)
    print("LOGIN AUTOVISION")
    print("=" * 70)

    driver.get(
        URL_LOGIN
    )

    wait = WebDriverWait(
        driver,
        TIMEOUT
    )

    campo_usuario = wait.until(
        EC.visibility_of_element_located(
            (
                By.ID,
                "usuario"
            )
        )
    )

    campo_usuario.clear()

    campo_usuario.send_keys(
        USUARIO
    )

    campo_senha = wait.until(
        EC.visibility_of_element_located(
            (
                By.ID,
                "senha"
            )
        )
    )

    campo_senha.clear()

    campo_senha.send_keys(
        SENHA
    )

    botao_login = wait.until(
        EC.element_to_be_clickable(
            (
                By.CSS_SELECTOR,
                "button[type='submit']"
            )
        )
    )

    driver.execute_script(
        "arguments[0].click();",
        botao_login
    )

    print(
        "Login enviado..."
    )

    try:

        wait.until(
            lambda d:
            len(
                d.find_elements(
                    By.ID,
                    "usuario"
                )
            ) == 0
        )

    except TimeoutException:

        raise RuntimeError(
            "Login não concluído."
        )

    print(
        "✓ Login concluído."
    )


# ============================================================
# PREENCHER DATETIME
# ============================================================

def preencher_datetime(
    driver,
    elemento_id,
    valor
):

    wait = WebDriverWait(
        driver,
        TIMEOUT
    )

    campo = wait.until(
        EC.presence_of_element_located(
            (
                By.ID,
                elemento_id
            )
        )
    )

    driver.execute_script(
        """
        const campo = arguments[0];
        const valor = arguments[1];

        campo.value = valor;

        campo.dispatchEvent(
            new Event('input', {
                bubbles: true
            })
        );

        campo.dispatchEvent(
            new Event('change', {
                bubbles: true
            })
        );

        campo.dispatchEvent(
            new Event('blur', {
                bubbles: true
            })
        );
        """,
        campo,
        valor
    )

    WebDriverWait(
        driver,
        TIMEOUT
    ).until(
        lambda d:
        d.find_element(
            By.ID,
            elemento_id
        ).get_attribute(
            "value"
        ) == valor
    )


# ============================================================
# CHECKBOX
# ============================================================

def selecionar_checkbox_por_texto(
    driver,
    texto_procurado
):

    texto_procurado = normalizar_texto(
        texto_procurado
    )

    labels = driver.find_elements(
        By.CSS_SELECTOR,
        "label.custom-control-label"
    )

    for label in labels:

        try:

            texto_label = normalizar_texto(
                label.text
            )

            if texto_procurado not in texto_label:
                continue

            input_element = None

            try:

                input_element = label.find_element(
                    By.CSS_SELECTOR,
                    "input"
                )

            except NoSuchElementException:
                pass

            if input_element is None:

                try:

                    input_element = label.find_element(
                        By.XPATH,
                        "./preceding-sibling::input[1]"
                    )

                except NoSuchElementException:
                    pass

            if input_element is None:

                try:

                    pai = label.find_element(
                        By.XPATH,
                        ".."
                    )

                    input_element = pai.find_element(
                        By.CSS_SELECTOR,
                        "input"
                    )

                except NoSuchElementException:
                    pass

            if input_element is not None:

                if not input_element.is_selected():

                    driver.execute_script(
                        "arguments[0].click();",
                        input_element
                    )

            else:

                driver.execute_script(
                    "arguments[0].click();",
                    label
                )

            print(
                f"✓ Campo selecionado: "
                f"{texto_procurado}"
            )

            return True

        except StaleElementReferenceException:

            continue

    print(
        f"⚠ Campo não encontrado: "
        f"{texto_procurado}"
    )

    return False


# ============================================================
# CONFIGURAR RELATÓRIO
# ============================================================

def configurar_relatorio_posicao(
    driver,
    inicio=None,
    fim=None
):

    wait = WebDriverWait(
        driver,
        TIMEOUT
    )

    if inicio is not None and fim is not None:

        preencher_datetime(
            driver,
            "data_inicial",
            inicio.strftime(
                "%Y-%m-%dT%H:%M"
            )
        )

        preencher_datetime(
            driver,
            "data_final",
            fim.strftime(
                "%Y-%m-%dT%H:%M"
            )
        )

    select_ponto = wait.until(
        EC.presence_of_element_located(
            (
                By.ID,
                "ponto"
            )
        )
    )

    Select(
        select_ponto
    ).select_by_value(
        "off"
    )

    print(
        "✓ Referência: Endereço"
    )

    selecionar_checkbox_por_texto(
        driver,
        "Velocidade"
    )

    selecionar_checkbox_por_texto(
        driver,
        "Motorista"
    )

    selecionar_checkbox_por_texto(
        driver,
        "Município"
    )


# ============================================================
# SELECIONAR PLACA
# ============================================================

def selecionar_placa_autovision(
    driver,
    placa
):

    placa = normalizar_placa(
        placa
    )

    wait = WebDriverWait(
        driver,
        TIMEOUT
    )

    print(
        f"Selecionando placa: {placa}"
    )

    campo_filtro = wait.until(
        EC.visibility_of_element_located(
            (
                By.CSS_SELECTOR,
                "input[name='q']"
            )
        )
    )

    campo_filtro.clear()

    campo_filtro.send_keys(
        placa
    )

    time.sleep(1)

    select_element = wait.until(
        EC.presence_of_element_located(
            (
                By.ID,
                "multiselect"
            )
        )
    )

    def encontrar_option(driver):

        try:

            elemento = driver.find_element(
                By.ID,
                "multiselect"
            )

            select = Select(
                elemento
            )

            for option in select.options:

                texto = normalizar_placa(
                    option.text
                )

                valor = normalizar_placa(
                    option.get_attribute(
                        "value"
                    )
                )

                if (
                    texto == placa
                    or
                    valor == placa
                ):

                    return True

        except Exception:
            pass

        return False

    wait.until(
        encontrar_option
    )

    encontrada = driver.execute_script(
        """
        const select = arguments[0];
        const placa = arguments[1];

        let encontrada = false;

        for (
            let i = 0;
            i < select.options.length;
            i++
        ) {

            const option = select.options[i];

            const texto =
                (option.text || "")
                .toUpperCase()
                .replace(/[^A-Z0-9]/g, "");

            const valor =
                (option.value || "")
                .toUpperCase()
                .replace(/[^A-Z0-9]/g, "");

            if (
                texto === placa ||
                valor === placa
            ) {

                option.selected = true;

                encontrada = true;

            } else {

                option.selected = false;
            }
        }

        select.dispatchEvent(
            new Event(
                "change",
                {
                    bubbles: true
                }
            )
        );

        return encontrada;
        """,
        select_element,
        placa
    )

    if not encontrada:

        raise RuntimeError(
            f"Placa {placa} não encontrada."
        )

    print(
        f"✓ {placa} encontrada."
    )

    botao_right_all = wait.until(
        EC.presence_of_element_located(
            (
                By.ID,
                "multiselect_rightAll"
            )
        )
    )

    driver.execute_script(
        """
        arguments[0].scrollIntoView({
            block: 'center'
        });

        arguments[0].click();
        """,
        botao_right_all
    )

    print(
        "✓ Placa movida para selecionados."
    )

    time.sleep(1)


# ============================================================
# LOCALIZAR BOTÃO RELATÓRIO
# ============================================================

def localizar_botao_relatorio_analitico(
    driver
):

    wait = WebDriverWait(
        driver,
        TIMEOUT
    )

    seletores = [

        (
            By.XPATH,
            "/html/body/section/form/div[2]/"
            "fieldset/div[1]/button[1]"
        ),

        (
            By.CSS_SELECTOR,
            "#formulario > div:nth-child(2) > "
            "fieldset > div.col-xs-7 > "
            "button.btn.submit-button"
        ),

        (
            By.CSS_SELECTOR,
            "button[onclick='relatorioPosicao(0,0);']"
        ),

        (
            By.XPATH,
            "//button[contains("
            "normalize-space(.),"
            "'Relatório analítico'"
            ")]"
        ),

        (
            By.XPATH,
            "//button[contains("
            "normalize-space(.),"
            "'Relatorio analitico'"
            ")]"
        ),
    ]

    for by, seletor in seletores:

        try:

            return wait.until(
                EC.presence_of_element_located(
                    (
                        by,
                        seletor
                    )
                )
            )

        except TimeoutException:
            continue

    try:

        botoes = driver.find_elements(
            By.CSS_SELECTOR,
            "button"
        )

        for botao in botoes:

            try:

                texto = normalizar_texto(
                    botao.text
                )

                onclick = (
                    botao.get_attribute(
                        "onclick"
                    )
                    or ""
                )

                if (
                    "RELATORIOPOSICAO"
                    in normalizar_texto(
                        onclick
                    )
                ):

                    return botao

                if (
                    "RELATORIO ANALITICO"
                    in texto
                ):

                    return botao

            except Exception:
                continue

    except Exception:
        pass

    return None


# ============================================================
# CLICAR RELATÓRIO
# ============================================================

def clicar_relatorio_analitico(
    driver
):

    abas_antes = set(
        driver.window_handles
    )

    botao = (
        localizar_botao_relatorio_analitico(
            driver
        )
    )

    if botao is None:

        raise RuntimeError(
            "Botão Relatório analítico "
            "não encontrado."
        )

    try:

        driver.execute_script(
            """
            arguments[0].scrollIntoView({
                block: 'center'
            });
            """,
            botao
        )

    except Exception:
        pass

    time.sleep(1)

    try:

        botao.click()

    except Exception:

        driver.execute_script(
            "arguments[0].click();",
            botao
        )

    try:

        WebDriverWait(
            driver,
            10
        ).until(
            lambda d:
            len(
                set(
                    d.window_handles
                )
                -
                abas_antes
            ) > 0
        )

    except TimeoutException:

        botao = (
            localizar_botao_relatorio_analitico(
                driver
            )
        )

        if botao is not None:

            try:

                driver.execute_script(
                    "arguments[0].click();",
                    botao
                )

            except Exception:
                pass

        try:

            novas = (
                set(
                    driver.window_handles
                )
                -
                abas_antes
            )

            if not novas:

                resultado = driver.execute_script(
                    """
                    if (
                        typeof relatorioPosicao ===
                        'function'
                    ) {

                        relatorioPosicao(0, 0);

                        return true;
                    }

                    return false;
                    """
                )

                if not resultado:

                    raise RuntimeError(
                        "relatorioPosicao indisponível."
                    )

        except Exception:
            pass

        WebDriverWait(
            driver,
            20
        ).until(
            lambda d:
            len(
                set(
                    d.window_handles
                )
                -
                abas_antes
            ) > 0
        )

    novas_abas = (
        set(
            driver.window_handles
        )
        -
        abas_antes
    )

    if not novas_abas:

        raise RuntimeError(
            "Nova aba não encontrada."
        )

    nova_aba = next(
        iter(novas_abas)
    )

    driver.switch_to.window(
        nova_aba
    )

    print(
        "✓ Nova aba aberta."
    )

    return nova_aba


# ============================================================
# AGUARDAR RELATÓRIO
# ============================================================

def aguardar_relatorio(
    driver
):

    wait = WebDriverWait(
        driver,
        TIMEOUT_RELATORIO
    )

    try:

        wait.until(
            lambda d:
            d.execute_script(
                "return document.readyState"
            ) == "complete"
        )

    except TimeoutException:
        pass

    wait.until(
        EC.presence_of_element_located(
            (
                By.TAG_NAME,
                "body"
            )
        )
    )

    def existe_tabela(driver):

        try:

            tabelas = driver.find_elements(
                By.TAG_NAME,
                "table"
            )

            for tabela in tabelas:

                linhas = tabela.find_elements(
                    By.CSS_SELECTOR,
                    "tbody tr"
                )

                if len(linhas) > 0:

                    return True

        except Exception:
            pass

        return False

    try:

        wait.until(
            existe_tabela
        )

        print(
            "✓ Tabela encontrada."
        )

        return True

    except TimeoutException:

        try:

            texto = normalizar_texto(
                driver.find_element(
                    By.TAG_NAME,
                    "body"
                ).text
            )

            if any(
                termo in texto
                for termo in [
                    "NENHUM REGISTRO",
                    "NENHUM DADO",
                    "SEM DADOS",
                ]
            ):

                return False

        except Exception:
            pass

        raise RuntimeError(
            "Tabela não encontrada."
        )


# ============================================================
# LOCALIZAR ÍCONE HTML
# ============================================================

def localizar_link_html(
    driver
):

    try:

        imagem = driver.find_element(
            By.ID,
            "btnHTML"
        )

        link = imagem.find_element(
            By.XPATH,
            "./ancestor::a[1]"
        )

        return link

    except Exception:
        pass

    try:

        links = driver.find_elements(
            By.CSS_SELECTOR,
            "a[download]"
        )

        for link in links:

            download = (
                link.get_attribute(
                    "download"
                )
                or ""
            )

            if (
                "RELATORIO_POSICAO"
                in normalizar_texto(
                    download
                )
            ):

                return link

    except Exception:
        pass

    return None


# ============================================================
# LISTAR HTMLS
# ============================================================

def listar_htmls():

    if not PASTA_HTML.exists():
        return set()

    return {
        arquivo.name
        for arquivo in PASTA_HTML.iterdir()
        if (
            arquivo.is_file()
            and
            arquivo.suffix.lower()
            in {
                ".html",
                ".htm"
            }
        )
    }


# ============================================================
# CLICAR HTML
# ============================================================

def clicar_icone_html(
    driver
):

    link = (
        localizar_link_html(
            driver
        )
    )

    if link is None:

        raise RuntimeError(
            "Ícone HTML não encontrado."
        )

    print(
        "✓ Ícone HTML encontrado."
    )

    driver.execute_script(
        """
        arguments[0].scrollIntoView({
            block: 'center'
        });
        """,
        link
    )

    time.sleep(1)

    try:

        link.click()

    except Exception:

        driver.execute_script(
            "arguments[0].click();",
            link
        )

    print(
        "✓ Clique no HTML realizado."
    )


# ============================================================
# ESPERAR DOWNLOAD HTML
# ============================================================

def esperar_download_html(arquivos_antes, timeout=None):

    if timeout is None:
        timeout = TIMEOUT_DOWNLOAD

    PASTA_HTML.mkdir(
        parents=True,
        exist_ok=True
    )

    limite = time.time() + timeout

    print("⬇ Aguardando download HTML...")

    while time.time() < limite:

        try:
            arquivos = list(
                PASTA_HTML.iterdir()
            )
        except Exception:
            arquivos = []

        # ----------------------------------------------------
        # Arquivos temporários
        # ----------------------------------------------------

        temporarios = []

        for arquivo in arquivos:

            if not arquivo.is_file():
                continue

            nome = arquivo.name.lower()

            if (
                nome.endswith(".crdownload")
                or
                nome.endswith(".tmp")
                or
                nome.endswith(".part")
            ):
                temporarios.append(arquivo)

        # ----------------------------------------------------
        # Novos HTMLs
        # ----------------------------------------------------

        novos = []

        for arquivo in arquivos:

            if not arquivo.is_file():
                continue

            if arquivo.name in arquivos_antes:
                continue

            if arquivo.suffix.lower() not in {
                ".html",
                ".htm"
            }:
                continue

            novos.append(
                arquivo
            )

        if novos and not temporarios:

            arquivo = max(
                novos,
                key=lambda x: x.stat().st_mtime
            )

            # ------------------------------------------------
            # Confirmar que terminou de gravar
            # ------------------------------------------------

            try:

                tamanho_anterior = (
                    arquivo.stat().st_size
                )

                time.sleep(1)

                if not arquivo.exists():
                    continue

                tamanho_atual = (
                    arquivo.stat().st_size
                )

                if (
                    tamanho_atual == tamanho_anterior
                    and
                    tamanho_atual > 0
                ):

                    print(
                        f"✓ HTML baixado: "
                        f"{arquivo.name}"
                    )

                    return arquivo

            except (
                FileNotFoundError,
                OSError
            ):
                continue

        time.sleep(0.5)

    # ========================================================
    # DIAGNÓSTICO
    # ========================================================

    print()
    print("!" * 70)
    print("❌ DOWNLOAD HTML NÃO CONFIRMADO")
    print("!" * 70)

    try:

        arquivos_finais = [
            arquivo.name
            for arquivo in PASTA_HTML.iterdir()
            if arquivo.is_file()
        ]

        print(
            "Arquivos encontrados na pasta HTML:"
        )

        if arquivos_finais:

            for nome in arquivos_finais[-20:]:
                print(
                    f"   - {nome}"
                )

        else:

            print(
                "   Nenhum arquivo encontrado."
            )

    except Exception as erro:

        print(
            f"⚠ Erro listando pasta HTML: {erro}"
        )

    return None


# ============================================================
# RENOMEAR HTML
# ============================================================

def renomear_html(
    arquivo,
    placa
):

    nome = (
        f"{nome_seguro(placa)}_"
        f"{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"
    )

    destino = (
        PASTA_HTML /
        nome
    )

    if destino.exists():

        destino.unlink()

    shutil.move(
        str(arquivo),
        str(destino)
    )

    return destino


# ============================================================
# PROCESSAR PLACA
# ============================================================

def processar_placa(
    driver,
    aba_principal,
    placa,
    inicio,
    fim
):

    aba_relatorio = None

    try:

        print()
        print("#" * 70)
        print(
            f"PROCESSANDO: {placa}"
        )
        print("#" * 70)

        driver.switch_to.window(
            aba_principal
        )

        driver.get(
            URL_RELATORIO_POSICAO
        )

        WebDriverWait(
            driver,
            TIMEOUT
        ).until(
            lambda d:
            d.execute_script(
                "return document.readyState"
            ) == "complete"
        )

        preencher_datetime(
            driver,
            "data_inicial",
            inicio.strftime(
                "%Y-%m-%dT%H:%M"
            )
        )

        preencher_datetime(
            driver,
            "data_final",
            fim.strftime(
                "%Y-%m-%dT%H:%M"
            )
        )

        print(
            "✓ Datas preenchidas."
        )

        configurar_relatorio_posicao(
            driver
        )

        selecionar_placa_autovision(
            driver,
            placa
        )

        aba_relatorio = (
            clicar_relatorio_analitico(
                driver
            )
        )

        possui_dados = (
            aguardar_relatorio(
                driver
            )
        )

        if not possui_dados:

            print(
                f"⚠ {placa}: sem registros."
            )

            return None

        arquivos_antes = listar_htmls()

        clicar_icone_html(
            driver
        )

        print(
            "⬇ Aguardando download HTML..."
        )

        arquivo = (
            esperar_download_html(
                arquivos_antes
            )
        )

        if arquivo is None:

            raise TimeoutException(
                "HTML não foi baixado."
            )

        arquivo_final = (
            renomear_html(
                arquivo,
                placa
            )
        )

        print(
            f"✓ HTML salvo: "
            f"{arquivo_final.name}"
        )

        return arquivo_final

    finally:

        try:

            if (
                aba_relatorio
                and
                aba_relatorio
                in driver.window_handles
            ):

                driver.close()

        except Exception:
            pass

        try:

            if (
                aba_principal
                in driver.window_handles
            ):

                driver.switch_to.window(
                    aba_principal
                )

        except Exception:
            pass


# ============================================================
# PROCESSAR TODAS AS PLACAS
# ============================================================

def processar_placas(
    driver,
    inicio,
    fim
):

    arquivos_html = []

    aba_principal = (
        driver.current_window_handle
    )

    total = len(
        PLACAS
    )

    for indice, placa in enumerate(
        PLACAS,
        start=1
    ):

        print()
        print(
            f"[{indice}/{total}]"
        )

        try:

            arquivo = (
                processar_placa(
                    driver,
                    aba_principal,
                    placa,
                    inicio,
                    fim
                )
            )

            if arquivo:

                arquivos_html.append(
                    arquivo
                )

        except Exception as erro:

            print()
            print("!" * 70)

            print(
                f"ERRO NA PLACA {placa}"
            )

            print(
                erro
            )

            print("!" * 70)

        time.sleep(
            INTERVALO_ENTRE_PLACAS
        )

    return arquivos_html




# ============================================================
# DATA / ARQUIVO
# ============================================================

def normalizar_data(valor):
    """Sempre retorna datetime.date, evitando .date() duplicado."""
    if valor is None:
        return None

    if isinstance(valor, datetime):
        return valor.date()

    from datetime import date as _date

    if isinstance(valor, _date):
        return valor

    try:
        convertido = pd.to_datetime(
            valor,
            dayfirst=True,
            errors="coerce"
        )

        if pd.isna(convertido):
            return None

        if isinstance(convertido, pd.Timestamp):
            return convertido.date()

        return convertido.date()

    except Exception:
        return None


def obter_data_arquivo(arquivo):
    match = PADRAO_ARQUIVO.search(arquivo.stem)

    if not match:
        return None

    try:
        return datetime.strptime(
            match.group("data"),
            "%Y%m%d"
        ).date()

    except ValueError:
        return None


def obter_placa_arquivo(arquivo):
    match = PADRAO_ARQUIVO.search(arquivo.stem)

    if match:
        return normalizar_placa(
            match.group("placa")
        )

    return normalizar_placa(
        arquivo.stem.split("_")[0]
    )


def listar_htmls_periodo(inicio, fim):

    inicio = normalizar_data(inicio)
    fim = normalizar_data(fim)

    arquivos = []

    for arquivo in PASTA_HTML.iterdir():

        if (
            not arquivo.is_file()
            or arquivo.suffix.lower()
            not in {".html", ".htm"}
        ):
            continue

        data = obter_data_arquivo(arquivo)

        # Arquivos sem data no padrão também são mantidos.
        if (
            data is None
            or (
                inicio <= data <= fim
            )
        ):
            arquivos.append(arquivo)

    return sorted(
        arquivos,
        key=lambda x: x.stat().st_mtime
    )


def nome_coluna_normalizado(nome):

    return re.sub(
        r"[^A-Z0-9]",
        "",
        normalizar_texto(nome)
    )


def encontrar_coluna(df, alternativas):

    mapa = {
        nome_coluna_normalizado(c): c
        for c in df.columns
    }

    for alternativa in alternativas:

        chave = nome_coluna_normalizado(
            alternativa
        )

        if chave in mapa:
            return mapa[chave]

    for chave, original in mapa.items():

        for alternativa in alternativas:

            alvo = nome_coluna_normalizado(
                alternativa
            )

            if alvo and (
                alvo in chave
                or chave in alvo
            ):
                return original

    return None


def converter_float(valor):

    if (
        valor is None
        or (
            isinstance(valor, float)
            and pd.isna(valor)
        )
    ):
        return 0.0

    texto = str(valor).strip().replace(
        " ",
        ""
    )

    if "," in texto and "." in texto:

        texto = (
            texto
            .replace(".", "")
            .replace(",", ".")
        )

    else:

        texto = texto.replace(
            ",",
            "."
        )

    try:
        return float(texto)

    except Exception:
        return 0.0


def limpar_motorista(valor):

    texto = normalizar_texto(valor)

    texto = re.sub(
        r"[^A-Z0-9 ]",
        " ",
        texto
    )

    texto = re.sub(
        r"\s+",
        " ",
        texto
    ).strip()

    invalidos = {
        "",
        "-",
        "--",
        "N A",
        "NA",
        "N/A",
        "NULL",
        "NONE",
        "SEM",
        "SEM MOTORISTA",
        "SEMMOTORISTA",
        "0",
    }

    if (
        texto in invalidos
        or len(texto) <= 2
    ):
        return ""

    return texto


def enderecos_compativeis(a, b):

    if not a or not b:
        return True

    a = normalizar_texto(a)
    b = normalizar_texto(b)

    a = re.sub(
        r"[^A-Z0-9 ]",
        " ",
        a
    )

    b = re.sub(
        r"[^A-Z0-9 ]",
        " ",
        b
    )

    a = re.sub(
        r"\s+",
        " ",
        a
    ).strip()

    b = re.sub(
        r"\s+",
        " ",
        b
    ).strip()

    return (
        a in b
        or b in a
    )


def ler_html_telemetria(arquivo):

    try:

        tabelas = pd.read_html(
            str(arquivo)
        )

    except Exception as erro:

        print(
            f"⚠ Erro lendo {arquivo.name}: {erro}"
        )

        return pd.DataFrame()

    registros = []

    placa = obter_placa_arquivo(
        arquivo
    )

    for df in tabelas:

        if df.empty:
            continue

        col_data = encontrar_coluna(
            df,
            [
                "Data Hora",
                "Data/Hora",
                "DataHora",
                "Data",
                "Hora",
            ]
        )

        col_ig = encontrar_coluna(
            df,
            [
                "IG",
                "Ig",
                "Ignição",
                "Ignicao",
                "Ignicao Motor",
            ]
        )

        col_vel = encontrar_coluna(
            df,
            [
                "Vel (Km/h)",
                "Vel(Km/h)",
                "Velocidade",
                "Vel",
                "Km/h",
            ]
        )

        col_end = encontrar_coluna(
            df,
            [
                "Endereço",
                "Endereco",
                "Localização",
                "Localizacao",
                "Local",
            ]
        )

        col_mot = encontrar_coluna(
            df,
            [
                "Motorista",
                "Condutor",
                "Driver",
            ]
        )

        col_lat = encontrar_coluna(
            df,
            [
                "Latitude",
                "Lat",
            ]
        )

        col_lon = encontrar_coluna(
            df,
            [
                "Longitude",
                "Long",
                "Lng",
            ]
        )

        col_mun = encontrar_coluna(
            df,
            [
                "Município",
                "Municipio",
            ]
        )

        if col_data is None:
            continue

        for _, row in df.iterrows():

            data_hora = pd.to_datetime(
                row.get(col_data),
                dayfirst=True,
                errors="coerce"
            )

            if pd.isna(data_hora):
                continue

            registros.append(
                {
                    "PLACA": placa,

                    "ARQUIVO_ORIGEM":
                        arquivo.name,

                    "DATA_HORA":
                        (
                            data_hora.to_pydatetime()
                            if isinstance(
                                data_hora,
                                pd.Timestamp
                            )
                            else data_hora
                        ),

                    "IG":
                        normalizar_texto(
                            row.get(
                                col_ig,
                                ""
                            )
                        ),

                    "VELOCIDADE":
                        converter_float(
                            row.get(
                                col_vel,
                                0
                            )
                        ),

                    "MOTORISTA":
                        str(
                            row.get(
                                col_mot,
                                ""
                            )
                            or ""
                        ).strip(),

                    "MOTORISTA_NORMALIZADO":
                        limpar_motorista(
                            row.get(
                                col_mot,
                                ""
                            )
                        ),

                    "ENDERECO":
                        str(
                            row.get(
                                col_end,
                                ""
                            )
                            or ""
                        ).strip(),

                    "MUNICIPIO":
                        str(
                            row.get(
                                col_mun,
                                ""
                            )
                            or ""
                        ).strip(),

                    "LATITUDE":
                        converter_float(
                            row.get(
                                col_lat,
                                0
                            )
                        ),

                    "LONGITUDE":
                        converter_float(
                            row.get(
                                col_lon,
                                0
                            )
                        ),
                }
            )

    if not registros:
        return pd.DataFrame()

    return (
        pd.DataFrame(registros)
        .drop_duplicates()
        .sort_values(
            [
                "PLACA",
                "DATA_HORA"
            ]
        )
        .reset_index(drop=True)
    )


# ============================================================
# FONTE DE ABASTECIMENTO
# ============================================================

def carregar_fonte_abastecimento(
    inicio,
    fim
):

    print(
        "\n" + "-" * 70
    )

    print(
        "CARREGANDO FONTE DE ABASTECIMENTO - GOOGLE SHEETS"
    )

    print(
        "-" * 70
    )

    print(
        f"URL: {URL_ABASTECIMENTO}"
    )

    try:

        resposta = requests.get(
            URL_ABASTECIMENTO,
            timeout=60,
            headers={
                "User-Agent":
                    "Mozilla/5.0 COMPESA-AuditoriaFrota/1.0"
            },
        )

        resposta.raise_for_status()

        conteudo = resposta.content

        try:

            texto = conteudo.decode(
                "utf-8-sig"
            )

        except UnicodeDecodeError:

            texto = conteudo.decode(
                "latin-1",
                errors="replace"
            )

        from io import StringIO

        try:

            df = pd.read_csv(
                StringIO(texto),
                sep=None,
                engine="python",
                dtype=object
            )

        except Exception:

            df = pd.read_csv(
                StringIO(texto),
                sep=",",
                dtype=object
            )

        if df.empty:

            print(
                "⚠ A fonte de abastecimento retornou 0 registros."
            )

            return pd.DataFrame()

        df.columns = [
            str(c).strip()
            for c in df.columns
        ]

        print(
            f"✓ Registros recebidos da fonte: {len(df)}"
        )

        print(
            "✓ Colunas:",
            ", ".join(
                map(
                    str,
                    df.columns.tolist()
                )
            )
        )

        placa_col = encontrar_coluna(
            df,
            [
                "PLACA",
                "Placa",
                "PLACA VEICULO",
                "PLACA VEÍCULO",
            ]
        )

        data_col = encontrar_coluna(
            df,
            [
                "DATA TRANSACAO",
                "DATA TRANSAÇÃO",
                "DATA",
                "DATA ABASTECIMENTO",
                "DATA ABAST.",
                "DATA/HORA",
                "DATA HORA",
            ]
        )

        if not placa_col or not data_col:

            print(
                "❌ A fonte de abastecimento não possui PLACA e/ou DATA TRANSACAO."
            )

            print(
                "PLACA encontrada:",
                placa_col
            )

            print(
                "DATA encontrada:",
                data_col
            )

            return pd.DataFrame()

        df[
            "PLACA_NORMALIZADA"
        ] = df[
            placa_col
        ].map(
            normalizar_placa
        )

        df[
            "DATA_ABAST"
        ] = pd.to_datetime(
            df[data_col],
            dayfirst=True,
            errors="coerce"
        )

        inicio_ts = pd.Timestamp(
            inicio
        )

        fim_ts = pd.Timestamp(
            fim
        )

        if (
            fim_ts.hour == 0
            and fim_ts.minute == 0
            and fim_ts.second == 0
        ):

            fim_ts = (
                fim_ts
                + pd.Timedelta(days=1)
                - pd.Timedelta(seconds=1)
            )

        antes = len(df)

        df = df[
            (df["PLACA_NORMALIZADA"] != "")
            &
            (df["DATA_ABAST"].notna())
            &
            (df["DATA_ABAST"] >= inicio_ts)
            &
            (df["DATA_ABAST"] <= fim_ts)
        ].copy()

        placas_validas = {
            normalizar_placa(p)
            for p in PLACAS
            if normalizar_placa(p)
        }

        if placas_validas:

            df = df[
                df[
                    "PLACA_NORMALIZADA"
                ].isin(
                    placas_validas
                )
            ].copy()

        df[
            "__ARQUIVO_ABAST"
        ] = (
            "Google Sheets - URL_ABASTECIMENTO"
        )

        df[
            "__ABA_ABAST"
        ] = "CSV publicado"

        print(
            f"✓ Registros após filtro de período/placa: "
            f"{len(df)} de {antes}"
        )

        if not df.empty:

            print(
                f"✓ Primeiro abastecimento: "
                f"{df['DATA_ABAST'].min()}"
            )

            print(
                f"✓ Último abastecimento:   "
                f"{df['DATA_ABAST'].max()}"
            )

        else:

            print(
                "⚠ Nenhum abastecimento encontrado para o período/placas."
            )

        return df.reset_index(
            drop=True
        )

    except Exception as erro:

        print(
            f"❌ ERRO AO CARREGAR FONTE DE ABASTECIMENTO: {erro}"
        )

        print(
            "   O processamento da telemetria continuará, "
            "mas os desvios de abastecimento não serão calculados."
        )

        return pd.DataFrame()


def extrair_dados_abastecimento(
    row
):

    def valor(cols):

        c = encontrar_coluna(
            pd.DataFrame([row]),
            cols
        )

        return (
            row.get(c, "")
            if c
            else ""
        )

    motorista = valor(
        [
            "NOME MOTORISTA",
            "MOTORISTA",
            "CONDUTOR",
            "NOME DO MOTORISTA",
            "NOME CONDUTOR",
        ]
    )

    endereco = valor(
        [
            "ENDERECO",
            "ENDEREÇO",
            "ENDEREÇO POSTO",
            "LOCALIZAÇÃO",
            "LOCALIZACAO",
        ]
    )

    bairro = valor(
        ["BAIRRO"]
    )

    cidade = valor(
        [
            "CIDADE",
            "MUNICIPIO",
            "MUNICÍPIO",
        ]
    )

    posto = valor(
        [
            "NOME POSTO",
            "POSTO",
            "NOME ESTABELECIMENTO",
            "ESTABELECIMENTO",
        ]
    )

    lat = converter_float(
        valor(
            [
                "LATITUDE",
                "LAT",
            ]
        )
    )

    lon = converter_float(
        valor(
            [
                "LONGITUDE",
                "LONG",
                "LNG",
            ]
        )
    )

    partes = [
        str(endereco or "").strip(),
        str(bairro or "").strip(),
        str(cidade or "").strip(),
    ]

    partes = [
        x
        for x in partes
        if (
            x
            and normalizar_texto(x)
            not in {
                "N/A",
                "NULL",
                "NAN",
                "-",
            }
        )
    ]

    return {
        "PLACA":
            normalizar_placa(
                row.get(
                    "PLACA_NORMALIZADA",
                    row.get(
                        "PLACA",
                        ""
                    )
                )
            ),

        "DATA_ABAST":
            row.get(
                "DATA_ABAST"
            ),

        "MOTORISTA_ABAST":
            str(
                motorista or ""
            ).strip(),

        "MOTORISTA_ABAST_NORMALIZADO":
            limpar_motorista(
                motorista
            ),

        "ENDERECO_ABAST":
            ", ".join(partes),

        "POSTO_ABAST":
            str(
                posto or ""
            ).strip(),

        "LATITUDE_ABAST":
            lat,

        "LONGITUDE_ABAST":
            lon,

        "ARQUIVO_FONTE_ABAST":
            row.get(
                "__ARQUIVO_ABAST",
                ""
            ),

        "ABA_FONTE_ABAST":
            row.get(
                "__ABA_ABAST",
                ""
            ),
    }


# ============================================================
# GEOCODIFICAÇÃO
# ============================================================

def geocodificar_endereco(
    endereco
):

    if not endereco:
        return None

    chave = (
        "addr",
        normalizar_texto(
            endereco
        )
    )

    if chave in _CACHE_LOCAIS:
        return _CACHE_LOCAIS[chave]

    global _ULTIMA_CONSULTA_OSM

    espera = (
        OSM_MIN_INTERVAL_S
        - (
            time.time()
            - _ULTIMA_CONSULTA_OSM
        )
    )

    if espera > 0:
        time.sleep(espera)

    try:

        resp = requests.get(
            "https://nominatim.openstreetmap.org/search",
            params={
                "q": endereco,
                "format": "jsonv2",
                "limit": 1,
                "countrycodes": "br",
            },
            headers={
                "User-Agent":
                    NOMINATIM_USER_AGENT
            },
            timeout=20,
        )

        resp.raise_for_status()

        _ULTIMA_CONSULTA_OSM = time.time()

        dados = resp.json()

        if dados:

            coords = (
                float(
                    dados[0]["lat"]
                ),
                float(
                    dados[0]["lon"]
                )
            )

            _CACHE_LOCAIS[
                chave
            ] = coords

            return coords

    except Exception as erro:

        print(
            f"⚠ Geocodificação indisponível para "
            f"'{endereco}': {erro}"
        )

    _CACHE_LOCAIS[
        chave
    ] = None

    return None


# ============================================================
# CRUZAMENTO ABASTECIMENTO X TELEMETRIA
# ============================================================

def cruzar_abastecimento_telemetria(
    telemetria,
    abastecimento
):

    if (
        telemetria.empty
        or abastecimento.empty
    ):
        return pd.DataFrame()

    eventos = []

    for _, fuelrow in abastecimento.iterrows():

        f = extrair_dados_abastecimento(
            fuelrow
        )

        placa = f["PLACA"]

        data_abast = pd.to_datetime(
            f["DATA_ABAST"],
            errors="coerce"
        )

        if (
            not placa
            or pd.isna(data_abast)
        ):
            continue

        candidatos = telemetria[
            telemetria["PLACA"]
            == placa
        ].copy()

        if candidatos.empty:
            continue

        candidatos[
            "DATA_HORA"
        ] = pd.to_datetime(
            candidatos["DATA_HORA"],
            errors="coerce"
        )

        candidatos = candidatos[
            candidatos["DATA_HORA"].notna()
        ].copy()

        candidatos = candidatos[
            candidatos[
                "DATA_HORA"
            ].dt.date
            == data_abast.date()
        ].copy()

        if candidatos.empty:
            continue

        candidatos[
            "DIF_MIN"
        ] = (
            candidatos["DATA_HORA"]
            - data_abast
        ).abs().dt.total_seconds() / 60.0

        candidatos = candidatos[
            candidatos["DIF_MIN"]
            <= TOLERANCIA_ABASTECIMENTO_MIN
        ].sort_values(
            "DIF_MIN"
        )

        if candidatos.empty:
            continue

        t = candidatos.iloc[0]

        dif = float(
            t["DIF_MIN"]
        )

        motorista_abast = (
            f["MOTORISTA_ABAST"]
        )

        motorista_t = str(
            t.get(
                "MOTORISTA",
                ""
            )
            or ""
        ).strip()

        motorista_abast_norm = (
            f[
                "MOTORISTA_ABAST_NORMALIZADO"
            ]
        )

        motorista_t_norm = (
            limpar_motorista(
                motorista_t
            )
        )

        endereco_abast = (
            f["ENDERECO_ABAST"]
        )

        endereco_t = str(
            t.get(
                "ENDERECO",
                ""
            )
            or ""
        ).strip()

        base = {
            **f,

            "DATA_HORA_TELEMETRIA":
                t["DATA_HORA"],

            "DIF_MIN":
                round(
                    dif,
                    2
                ),

            "Motorista abastecimento":
                motorista_abast,

            "Motoristas Telemetria":
                motorista_t,

            "ENDERECO_TELEMETRIA":
                endereco_t,
        }

        # --------------------------------------------------------
        # MOTORISTA NÃO IDENTIFICADO
        # --------------------------------------------------------

        if not motorista_t_norm:

            eventos.append(
                {
                    **base,

                    "TIPO_DESVIO":
                        "Motorista não identificado",

                    "CRITICIDADE":
                        "MEDIA",

                    "DETALHE":
                        (
                            f"Motorista abastecimento: "
                            f"{motorista_abast or 'NÃO INFORMADO'} | "
                            f"Motoristas Telemetria: "
                            f"NÃO INFORMADO | "
                            f"Data/hora: "
                            f"{data_abast.strftime('%d/%m/%Y %H:%M:%S')} | "
                            f"Diferença: "
                            f"{dif:.2f} min"
                        ),
                }
            )

        # --------------------------------------------------------
        # MOTORISTA DIFERENTE
        # --------------------------------------------------------

        elif (
            motorista_abast_norm
            and motorista_t_norm
            and motorista_abast_norm
            != motorista_t_norm
        ):

            eventos.append(
                {
                    **base,

                    "TIPO_DESVIO":
                        "MOTORISTA DIFERENTE",

                    "CRITICIDADE":
                        "ALTA",

                    "DETALHE":
                        (
                            f"Motorista abastecimento: "
                            f"{motorista_abast} | "
                            f"Motoristas Telemetria: "
                            f"{motorista_t} | "
                            f"Data/hora: "
                            f"{data_abast.strftime('%d/%m/%Y %H:%M:%S')} | "
                            f"Diferença: "
                            f"{dif:.2f} min"
                        ),
                }
            )

        # --------------------------------------------------------
        # ENDEREÇO DIVERGENTE
        # --------------------------------------------------------

        endereco_div = False
        distancia_end = None

        latf = f[
            "LATITUDE_ABAST"
        ]

        lonf = f[
            "LONGITUDE_ABAST"
        ]

        latt = converter_float(
            t.get(
                "LATITUDE",
                0
            )
        )

        lont = converter_float(
            t.get(
                "LONGITUDE",
                0
            )
        )

        if (
            not (latf and lonf)
            and endereco_abast
        ):

            geo = geocodificar_endereco(
                endereco_abast
            )

            if geo:
                latf, lonf = geo

        if (
            not (latt and lont)
            and endereco_t
        ):

            geo = geocodificar_endereco(
                endereco_t
            )

            if geo:
                latt, lont = geo

        if (
            latf
            and lonf
            and latt
            and lont
        ):

            distancia_end = (
                distancia_metros(
                    latf,
                    lonf,
                    latt,
                    lont
                )
            )

            endereco_div = (
                distancia_end
                > RAIO_DIVERGENCIA_ENDERECO_METROS
            )

        else:

            endereco_div = not enderecos_compativeis(
                endereco_abast,
                endereco_t
            )

        if endereco_div:

            detalhe = (
                f"Abastecimento: "
                f"{endereco_abast or f['POSTO_ABAST'] or 'NÃO INFORMADO'} | "
                f"Telemetria: "
                f"{endereco_t or 'NÃO INFORMADO'} | "
                f"Data/hora: "
                f"{data_abast.strftime('%d/%m/%Y %H:%M:%S')} | "
                f"Diferença: "
                f"{dif:.2f} min"
            )

            if distancia_end is not None:

                detalhe += (
                    f" | Distância: "
                    f"{distancia_end:.1f} m"
                )

            eventos.append(
                {
                    **base,

                    "TIPO_DESVIO":
                        "ENDEREÇO DIVERGENTE",

                    "CRITICIDADE":
                        (
                            "ALTA"
                            if (
                                distancia_end
                                is not None
                                and distancia_end > 500
                            )
                            else "MEDIA"
                        ),

                    "DISTANCIA_ENDERECOS_M":
                        (
                            round(
                                distancia_end,
                                1
                            )
                            if distancia_end
                            is not None
                            else None
                        ),

                    "DETALHE":
                        detalhe,
                }
            )

        # --------------------------------------------------------
        # MOTOR LIGADO NO ABASTECIMENTO
        # --------------------------------------------------------

        if (
            dif
            <= TOLERANCIA_IGNICAO_ABASTECIMENTO_MIN
            and motor_ligado_parado(t)
        ):

            eventos.append(
                {
                    **base,

                    "TIPO_DESVIO":
                        "MOTOR LIGADO NO ABASTECIMENTO",

                    "CRITICIDADE":
                        "ALTA",

                    "DETALHE":
                        (
                            "IG = L e Vel (Km/h) = 0,0 "
                            "no registro de telemetria "
                            "correspondente ao abastecimento. "
                            f"Diferença: {dif:.2f} min."
                        ),
                }
            )

    if not eventos:
        return pd.DataFrame()

    return pd.DataFrame(
        eventos
    )


def identificar_desvios_local(
    df
):

    """
    Mantida por compatibilidade.

    Os desvios de motorista/endereço/motor ligado
    são gerados somente no cruzamento abastecimento
    x telemetria.
    """

    return pd.DataFrame()


def motor_ligado_parado(
    row
):

    return (
        normalizar_texto(
            row["IG"]
        ) == "L"
        and
        abs(
            float(
                row["VELOCIDADE"]
            )
        ) < 0.01
    )


def distancia_metros(
    lat1,
    lon1,
    lat2,
    lon2
):

    raio = 6371000.0

    p1 = math.radians(
        float(lat1)
    )

    p2 = math.radians(
        float(lat2)
    )

    dp = math.radians(
        float(lat2)
        - float(lat1)
    )

    dl = math.radians(
        float(lon2)
        - float(lon1)
    )

    a = (
        math.sin(dp / 2) ** 2
        +
        math.cos(p1)
        * math.cos(p2)
        * math.sin(dl / 2) ** 2
    )

    return (
        raio
        * 2
        * math.atan2(
            math.sqrt(a),
            math.sqrt(1 - a)
        )
    )


# ============================================================
# PARADAS COM MOTOR LIGADO
# ============================================================

def identificar_paradas_motor_ligado(
    df
):

    eventos = []

    if df.empty:
        return pd.DataFrame()

    for placa, grupo in df.groupby(
        "PLACA",
        sort=False
    ):

        grupo = (
            grupo
            .sort_values("DATA_HORA")
            .reset_index(drop=True)
        )

        inicio = None
        ultimo = None

        for i, row in grupo.iterrows():

            valido = motor_ligado_parado(
                row
            )

            if valido:

                if inicio is None:

                    inicio = i
                    ultimo = i

                    continue

                anterior = grupo.iloc[
                    ultimo
                ]

                delta = (
                    row["DATA_HORA"]
                    - anterior["DATA_HORA"]
                ).total_seconds() / 60.0

                mesma_posicao = True

                if (
                    all(
                        float(
                            anterior.get(
                                c,
                                0
                            )
                        ) != 0
                        for c in [
                            "LATITUDE",
                            "LONGITUDE"
                        ]
                    )
                    and
                    all(
                        float(
                            row.get(
                                c,
                                0
                            )
                        ) != 0
                        for c in [
                            "LATITUDE",
                            "LONGITUDE"
                        ]
                    )
                ):

                    mesma_posicao = (
                        distancia_metros(
                            anterior["LATITUDE"],
                            anterior["LONGITUDE"],
                            row["LATITUDE"],
                            row["LONGITUDE"]
                        )
                        <= RAIO_PARADA_METROS
                    )

                if (
                    delta
                    <= GAP_MAX_MINUTOS
                    and mesma_posicao
                ):

                    ultimo = i

                else:

                    evento = montar_parada(
                        grupo,
                        inicio,
                        ultimo
                    )

                    if evento:
                        eventos.append(
                            evento
                        )

                    inicio = i
                    ultimo = i

            elif inicio is not None:

                evento = montar_parada(
                    grupo,
                    inicio,
                    ultimo
                )

                if evento:
                    eventos.append(
                        evento
                    )

                inicio = None
                ultimo = None

        if inicio is not None:

            evento = montar_parada(
                grupo,
                inicio,
                ultimo
            )

            if evento:
                eventos.append(
                    evento
                )

    return pd.DataFrame(
        eventos
    )


def montar_parada(
    grupo,
    inicio,
    fim
):

    primeiro = grupo.iloc[
        inicio
    ]

    ultimo = grupo.iloc[
        fim
    ]

    minutos = (
        ultimo["DATA_HORA"]
        - primeiro["DATA_HORA"]
    ).total_seconds() / 60.0

    if minutos < TEMPO_MINIMO_PARADA:
        return None

    return {
        "PLACA":
            primeiro["PLACA"],

        "ARQUIVO_ORIGEM":
            primeiro["ARQUIVO_ORIGEM"],

        "INICIO":
            primeiro["DATA_HORA"],

        "FIM":
            ultimo["DATA_HORA"],

        "DURACAO_MINUTOS":
            round(
                minutos,
                1
            ),

        "MOTORISTA":
            primeiro["MOTORISTA"],

        "ENDERECO":
            (
                primeiro["ENDERECO"]
                or ultimo["ENDERECO"]
            ),

        "MUNICIPIO":
            primeiro["MUNICIPIO"],

        "LATITUDE":
            primeiro["LATITUDE"],

        "LONGITUDE":
            primeiro["LONGITUDE"],

        "TIPO_DESVIO":
            "MOTOR LIGADO - VEÍCULO PARADO",

        "LOCAL_TIPO":
            "",

        "LOCAL_NOME":
            "",

        "DISTANCIA_LOCAL_M":
            pd.NA,

        "CRITICIDADE":
            "",

        "DETALHE":
            "",
    }


# ============================================================
# LOCAL / OSM
# ============================================================

def classificar_tipo_local(
    tags
):

    for chave in OSM_TAG_KEYS:

        if chave not in tags:
            continue

        valor = str(
            tags.get(
                chave,
                ""
            )
        ).strip().lower()

        nivel = OSM_CRITICIDADE_MAP.get(
            (
                chave,
                valor
            )
        )

        if nivel:

            return (
                nivel,
                f"{chave}: {valor}"
            )

    return (
        "C",
        "Outro"
    )


def consultar_osm_proximo(
    lat,
    lon
):

    global _ULTIMA_CONSULTA_OSM

    query = f"""
    [out:json][timeout:25];
    (
      nwr(around:{RAIO_BUSCA_POI_METROS},{lat},{lon})[amenity];
      nwr(around:{RAIO_BUSCA_POI_METROS},{lat},{lon})[shop];
      nwr(around:{RAIO_BUSCA_POI_METROS},{lat},{lon})[tourism];
      nwr(around:{RAIO_BUSCA_POI_METROS},{lat},{lon})[leisure];
      nwr(around:{RAIO_BUSCA_POI_METROS},{lat},{lon})[natural];
      nwr(around:{RAIO_BUSCA_POI_METROS},{lat},{lon})[aeroway];
    );
    out center tags;
    """

    ultimo_erro = None

    for endpoint in OVERPASS_URLS:

        try:

            agora = time.time()

            espera = (
                OSM_MIN_INTERVAL_S
                - (
                    agora
                    - _ULTIMA_CONSULTA_OSM
                )
            )

            if espera > 0:
                time.sleep(espera)

            resposta = requests.post(
                endpoint,
                data=query,
                headers={
                    "User-Agent":
                        NOMINATIM_USER_AGENT
                },
                timeout=40,
            )

            resposta.raise_for_status()

            _ULTIMA_CONSULTA_OSM = (
                time.time()
            )

            locais = []

            for item in resposta.json().get(
                "elements",
                []
            ):

                tags = item.get(
                    "tags",
                    {}
                )

                criticidade, categoria = (
                    classificar_tipo_local(
                        tags
                    )
                )

                centro = item.get(
                    "center",
                    {}
                )

                lat2 = item.get(
                    "lat",
                    centro.get(
                        "lat",
                        0
                    )
                )

                lon2 = item.get(
                    "lon",
                    centro.get(
                        "lon",
                        0
                    )
                )

                if not lat2 or not lon2:
                    continue

                locais.append(
                    {
                        "nome":
                            tags.get(
                                "name",
                                ""
                            ),

                        "categoria":
                            categoria,

                        "criticidade":
                            criticidade,

                        "lat":
                            lat2,

                        "lon":
                            lon2,
                    }
                )

            return locais

        except Exception as erro:

            ultimo_erro = erro
            continue

    print(
        "⚠ Nenhum servidor Overpass respondeu. "
        f"Análise continuará sem POI. "
        f"Último erro: {ultimo_erro}"
    )

    return []


# ============================================================
# GOOGLE PLACES
# ============================================================

def consultar_google_places(
    lat,
    lon
):

    if not GOOGLE_PLACES_API_KEY:
        return []

    url = (
        "https://places.googleapis.com/v1/places:searchNearby"
    )

    headers = {
        "Content-Type":
            "application/json",

        "X-Goog-Api-Key":
            GOOGLE_PLACES_API_KEY,

        "X-Goog-FieldMask":
            (
                "places.displayName,"
                "places.primaryType,"
                "places.location,"
                "places.types"
            ),
    }

    payload = {
        "includedTypes": [
            "bar",
            "night_club",
            "restaurant",
            "hospital",
            "shopping_mall",
            "supermarket",
            "lodging",
            "cafe",
            "clinic",
            "gym",
            "school",
            "gas_station",
            "parking_lot",
            "convenience_store",
            "department_store",
        ],

        "maxResultCount": 10,

        "locationRestriction": {
            "circle": {
                "center": {
                    "latitude":
                        lat,

                    "longitude":
                        lon,
                },

                "radius":
                    RAIO_BUSCA_POI_METROS,
            }
        },
    }

    try:

        resposta = requests.post(
            url,
            headers=headers,
            json=payload,
            timeout=20,
        )

        resposta.raise_for_status()

        locais = []

        for item in resposta.json().get(
            "places",
            []
        ):

            tipo = item.get(
                "primaryType",
                ""
            )

            nivel = CRITICIDADE_MAP.get(
                tipo,
                "C"
            )

            nomes = {
                "bar":
                    "Bar",

                "night_club":
                    "Vida noturna",

                "lodging":
                    "Hospedagem",

                "shopping_mall":
                    "Shopping",

                "hospital":
                    "Hospital",

                "restaurant":
                    "Restaurante",

                "supermarket":
                    "Supermercado",

                "cafe":
                    "Café",

                "gym":
                    "Academia",

                "school":
                    "Escola",

                "doctor":
                    "Atendimento médico",

                "gas_station":
                    "Posto",

                "parking":
                    "Estacionamento",
            }

            criticidade = nivel

            categoria = nomes.get(
                tipo,
                tipo.replace(
                    "_",
                    " "
                ).title()
            )

            loc = item.get(
                "location",
                {}
            )

            locais.append(
                {
                    "nome":
                        item.get(
                            "displayName",
                            {}
                        ).get(
                            "text",
                            ""
                        ),

                    "categoria":
                        categoria,

                    "criticidade":
                        criticidade,

                    "lat":
                        loc.get(
                            "latitude",
                            0
                        ),

                    "lon":
                        loc.get(
                            "longitude",
                            0
                        ),
                }
            )

        return locais

    except Exception as erro:

        print(
            f"⚠ Google Places indisponível: {erro}"
        )

        return []


# ============================================================
# VALIDAR LOCAL DAS PARADAS
# ============================================================

def validar_local_paradas(
    df
):

    if df.empty:
        return df

    resultado = df.copy()

    resultado[
        "DISTANCIA_LOCAL_M"
    ] = pd.to_numeric(
        resultado[
            "DISTANCIA_LOCAL_M"
        ],
        errors="coerce"
    ).astype("Float64")

    peso = CRITICIDADE_PESO

    for idx, row in resultado.iterrows():

        lat = converter_float(
            row.get(
                "LATITUDE",
                0
            )
        )

        lon = converter_float(
            row.get(
                "LONGITUDE",
                0
            )
        )

        duracao = converter_float(
            row.get(
                "DURACAO_MINUTOS",
                0
            )
        )

        melhor = None

        if lat and lon:

            chave = (
                round(lat, 4),
                round(lon, 4)
            )

            if chave not in _CACHE_LOCAIS:

                locais = []

                if GOOGLE_PLACES_API_KEY:

                    print(
                        "🔎 Google Places API: "
                        "consultando primeiro..."
                    )

                    locais = (
                        consultar_google_places(
                            lat,
                            lon
                        )
                    )

                if not locais:

                    print(
                        "⚠ Google Places sem "
                        "resultado/disponível. "
                        "Usando OSM/Overpass "
                        "como fallback..."
                    )

                    locais = (
                        consultar_osm_proximo(
                            lat,
                            lon
                        )
                    )

                _CACHE_LOCAIS[
                    chave
                ] = locais

            else:

                locais = _CACHE_LOCAIS[
                    chave
                ]

            for local in locais:

                distancia = distancia_metros(
                    lat,
                    lon,
                    local["lat"],
                    local["lon"]
                )

                candidato = {
                    **local,
                    "distancia":
                        round(
                            distancia,
                            1
                        )
                }

                if (
                    distancia
                    > RAIO_BUSCA_POI_METROS
                ):
                    continue

                if (
                    melhor is None
                    or
                    peso[
                        candidato[
                            "criticidade"
                        ]
                    ]
                    >
                    peso[
                        melhor[
                            "criticidade"
                        ]
                    ]
                    or
                    (
                        peso[
                            candidato[
                                "criticidade"
                            ]
                        ]
                        ==
                        peso[
                            melhor[
                                "criticidade"
                            ]
                        ]
                        and
                        distancia
                        <
                        melhor[
                            "distancia"
                        ]
                    )
                ):

                    melhor = candidato

        if melhor:

            criticidade = (
                melhor[
                    "criticidade"
                ]
            )

            resultado.at[
                idx,
                "LOCAL_TIPO"
            ] = melhor[
                "categoria"
            ]

            resultado.at[
                idx,
                "LOCAL_NOME"
            ] = melhor[
                "nome"
            ]

            resultado.at[
                idx,
                "DISTANCIA_LOCAL_M"
            ] = float(
                melhor[
                    "distancia"
                ]
            )

            resultado.at[
                idx,
                "CRITICIDADE"
            ] = criticidade

            resultado.at[
                idx,
                "DETALHE"
            ] = (
                f"Motor ligado e velocidade "
                f"0,0 km/h por {duracao:.1f} min; "
                f"local próximo: "
                f"{melhor['nome'] or 'sem nome'} "
                f"({melhor['categoria']}), "
                f"{melhor['distancia']:.0f} m."
            )

        else:

            resultado.at[
                idx,
                "CRITICIDADE"
            ] = "C"

            resultado.at[
                idx,
                "DETALHE"
            ] = (
                f"Motor ligado e velocidade "
                f"0,0 km/h por {duracao:.1f} min; "
                f"nenhum estabelecimento "
                f"classificado encontrado em "
                f"{RAIO_BUSCA_POI_METROS} m."
            )

    return resultado


# ============================================================
# ANALISAR RELATÓRIOS SALVOS
# ============================================================

def analisar_relatorios_salvos(
    arquivos_html,
    inicio=None,
    fim=None
):

    dados = []

    for arquivo in arquivos_html:

        df = ler_html_telemetria(
            arquivo
        )

        if not df.empty:
            dados.append(df)

    if not dados:

        return (
            pd.DataFrame(),
            pd.DataFrame(),
            pd.DataFrame(),
            pd.DataFrame()
        )

    telemetria = pd.concat(
        dados,
        ignore_index=True,
        sort=False
    )

    telemetria = (
        telemetria
        .drop_duplicates()
        .sort_values(
            [
                "PLACA",
                "DATA_HORA"
            ]
        )
        .reset_index(drop=True)
    )

    desvios = pd.DataFrame()

    paradas = (
        identificar_paradas_motor_ligado(
            telemetria
        )
    )

    if not paradas.empty:

        paradas = validar_local_paradas(
            paradas
        )

    if (
        inicio is not None
        and fim is not None
    ):

        abastecimento = (
            carregar_fonte_abastecimento(
                inicio,
                fim
            )
        )

    else:

        abastecimento = pd.DataFrame()

    desvios_abast = (
        cruzar_abastecimento_telemetria(
            telemetria,
            abastecimento
        )
    )

    return (
        telemetria,
        desvios,
        paradas,
        desvios_abast
    )


# ============================================================
# GERAR EXCEL DE AUDITORIA
# ============================================================

def gerar_excel_auditoria(
    telemetria,
    desvios,
    paradas,
    desvios_abast=None
):

    if desvios_abast is None:
        desvios_abast = pd.DataFrame()

    if not desvios_abast.empty:

        desvios = pd.concat(
            [
                desvios,
                desvios_abast
            ],
            ignore_index=True,
            sort=False
        )

    resumo = []

    if not desvios.empty:

        for tipo, qtd in (
            desvios[
                "TIPO_DESVIO"
            ].value_counts().items()
        ):

            resumo.append(
                {
                    "CATEGORIA":
                        tipo,

                    "QUANTIDADE":
                        int(qtd),
                }
            )

    if not paradas.empty:

        for tipo, qtd in (
            paradas[
                "CRITICIDADE"
            ].value_counts().items()
        ):

            resumo.append(
                {
                    "CATEGORIA":
                        f"PARADA - {tipo}",

                    "QUANTIDADE":
                        int(qtd),
                }
            )

    df_resumo = pd.DataFrame(
        resumo,
        columns=[
            "CATEGORIA",
            "QUANTIDADE"
        ]
    )

    with pd.ExcelWriter(
        ARQUIVO_AUDITORIA,
        engine="openpyxl"
    ) as writer:

        telemetria.to_excel(
            writer,
            sheet_name="Telemetria",
            index=False
        )

        desvios.to_excel(
            writer,
            sheet_name="Desvios",
            index=False
        )

        if not desvios_abast.empty:

            desvios_abast.to_excel(
                writer,
                sheet_name=
                    "Abastecimento_x_Telemetria",
                index=False
            )

        paradas.to_excel(
            writer,
            sheet_name=
                "Paradas_Motor_Ligado",
            index=False
        )

        df_resumo.to_excel(
            writer,
            sheet_name="Resumo",
            index=False
        )

    wb = load_workbook(
        ARQUIVO_AUDITORIA
    )

    for ws in wb.worksheets:

        ws.freeze_panes = "A2"

        if ws.max_row >= 1:

            ws.auto_filter.ref = (
                ws.dimensions
            )

        for cell in ws[1]:

            cell.font = Font(
                bold=True
            )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

        for col in ws.columns:

            maior = max(
                10,
                max(
                    len(
                        str(
                            c.value
                            or ""
                        )
                    )
                    for c in col
                ) + 2
            )

            ws.column_dimensions[
                get_column_letter(
                    col[0].column
                )
            ].width = min(
                maior,
                60
            )

    wb.save(
        ARQUIVO_AUDITORIA
    )

    return ARQUIVO_AUDITORIA


# ============================================================
# EXECUTAR AUDITORIA
# ============================================================

def executar_auditoria(
    arquivos_html,
    inicio,
    fim
):

    print(
        "\n" + "=" * 70
    )

    print(
        "AUDITORIA LOCAL - DESVIOS DE TELEMETRIA"
    )

    print(
        "=" * 70
    )

    print(
        f"Período: "
        f"{inicio.strftime('%d/%m/%Y')} "
        f"até "
        f"{fim.strftime('%d/%m/%Y')}"
    )

    print(
        f"HTMLs analisados: "
        f"{len(arquivos_html)}"
    )

    print(
        f"Fonte de abastecimento "
        f"(CSV publicado): "
        f"{URL_ABASTECIMENTO}"
    )

    print(
        f"Planilha origem (edição): "
        f"{URL_ABASTECIMENTO_EDIT}"
    )

    print(
        "Regra de cruzamento: "
        "mesma placa + mesmo dia + "
        "diferença máxima de 3 minutos"
    )

    (
        telemetria,
        desvios,
        paradas,
        desvios_abast
    ) = analisar_relatorios_salvos(
        arquivos_html,
        inicio,
        fim
    )

    print(
        f"✓ Fonte Google Sheets: "
        f"{len(desvios_abast)} "
        f"eventos de abastecimento x telemetria"
    )

    if telemetria.empty:

        print(
            "⚠ Nenhum registro de "
            "telemetria foi extraído."
        )

        return None

    arquivo = gerar_excel_auditoria(
        telemetria,
        desvios,
        paradas,
        desvios_abast
    )

    print(
        f"✓ Registros: "
        f"{len(telemetria)}"
    )

    print(
        f"✓ Desvios: "
        f"{len(desvios)}"
    )

    print(
        f"✓ Paradas motor ligado: "
        f"{len(paradas)}"
    )

    print(
        f"✓ Desvios abastecimento "
        f"x telemetria: "
        f"{len(desvios_abast)}"
    )

    print(
        f"✓ Excel: "
        f"{arquivo.resolve()}"
    )

    return arquivo


# ============================================================
# LOGIN NEXUS
# ============================================================

def login_nexus(
    driver
):

    print()

    print(
        "=" * 70
    )

    print(
        "NEXUS FROTA BI"
    )

    print(
        "=" * 70
    )

    driver.get(
        URL_NEXUS
    )

    time.sleep(3)

    wait = WebDriverWait(
        driver,
        TIMEOUT
    )

    try:

        botao = wait.until(
            EC.element_to_be_clickable(
                (
                    By.XPATH,
                    "//button[contains("
                    "normalize-space(.),"
                    "'Acesso Restrito'"
                    ")]"
                )
            )
        )

        driver.execute_script(
            "arguments[0].click();",
            botao
        )

        print(
            "✓ Acesso Restrito."
        )

        time.sleep(2)

    except TimeoutException:

        pass

    campo_email = None

    try:

        campo_email = wait.until(
            EC.visibility_of_element_located(
                (
                    By.CSS_SELECTOR,
                    "input[type='email']"
                )
            )
        )

    except TimeoutException:

        pass

    campo_senha = None

    try:

        campo_senha = wait.until(
            EC.visibility_of_element_located(
                (
                    By.CSS_SELECTOR,
                    "input[type='password']"
                )
            )
        )

    except TimeoutException:

        pass

    if (
        campo_email is None
        or campo_senha is None
    ):

        print(
            "⚠ Campos não localizados."
        )

        input(
            "Realize o login manualmente "
            "e pressione ENTER..."
        )

        return

    campo_email.clear()

    campo_email.send_keys(
        NEXUS_EMAIL
    )

    campo_senha.clear()

    if not NEXUS_SENHA:

        print(
            "⚠ NEXUS_SENHA não configurada."
        )

        input(
            "Digite a senha manualmente "
            "e pressione ENTER..."
        )

        return

    campo_senha.send_keys(
        NEXUS_SENHA
    )

    botao_entrar = wait.until(
        EC.element_to_be_clickable(
            (
                By.CSS_SELECTOR,
                "button[type='submit']"
            )
        )
    )

    driver.execute_script(
        "arguments[0].click();",
        botao_entrar
    )

    print(
        "✓ Login enviado."
    )

    time.sleep(5)


# ============================================================
# CLICAR NEXUS FROTA BI
# ============================================================

def clicar_nexus_frota_bi(
    driver
):

    wait = WebDriverWait(
        driver,
        TIMEOUT
    )

    try:

        elemento = wait.until(
            EC.presence_of_element_located(
                (
                    By.XPATH,
                    "//*[contains("
                    "normalize-space(.),"
                    "'Nexus Frota BI'"
                    ")]"
                )
            )
        )

        try:

            pai = elemento.find_element(
                By.XPATH,
                "./ancestor::*["
                "self::button or self::a"
                "][1]"
            )

            elemento = pai

        except Exception:

            pass

        driver.execute_script(
            "arguments[0].click();",
            elemento
        )

        print(
            "✓ Nexus Frota BI."
        )

        time.sleep(4)

        return True

    except Exception as erro:

        print(
            f"⚠ Erro Nexus Frota BI: "
            f"{erro}"
        )

        return False


# ============================================================
# MONITORAMENTO E ANÁLISE
# ============================================================

def acessar_monitoramento(
    driver
):

    wait = WebDriverWait(
        driver,
        TIMEOUT
    )

    try:

        elemento = wait.until(
            EC.element_to_be_clickable(
                (
                    By.XPATH,
                    "//button[contains("
                    "normalize-space(.),"
                    "'Monitoramento e Análise'"
                    ")]"
                )
            )
        )

    except TimeoutException:

        elemento = wait.until(
            EC.element_to_be_clickable(
                (
                    By.XPATH,
                    "//button[contains("
                    "normalize-space(.),"
                    "'Monitoramento e Analise'"
                    ")]"
                )
            )
        )

    driver.execute_script(
        "arguments[0].click();",
        elemento
    )

    print(
        "✓ Monitoramento e Análise."
    )

    time.sleep(3)


# ============================================================
# ABA ABASTECIMENTO X TELEMETRIA
# ============================================================

def acessar_abastecimento_telemetria(
    driver
):

    wait = WebDriverWait(
        driver,
        TIMEOUT
    )

    elemento = wait.until(
        EC.element_to_be_clickable(
            (
                By.XPATH,
                "//button[contains("
                "normalize-space(.),"
                "'Abastecimento x Telemetria'"
                ")]"
            )
        )
    )

    driver.execute_script(
        "arguments[0].click();",
        elemento
    )

    print(
        "✓ Abastecimento x Telemetria."
    )

    time.sleep(3)


# ============================================================
# INPUT HTML
# ============================================================

def localizar_input_html(
    driver
):

    seletores = [
        'input[type="file"][accept=".html"]',
        'input[type="file"][accept=".HTML"]',
        'input[type="file"]',
    ]

    for seletor in seletores:

        try:

            elementos = driver.find_elements(
                By.CSS_SELECTOR,
                seletor
            )

            for elemento in elementos:

                if elemento.is_enabled():

                    return elemento

        except Exception:

            continue

    return None


# ============================================================
# ENVIAR HTML
# ============================================================

def enviar_html(
    driver,
    arquivo_html
):

    campo = None

    limite = (
        time.time()
        + TIMEOUT
    )

    while time.time() < limite:

        campo = localizar_input_html(
            driver
        )

        if campo:
            break

        time.sleep(0.5)

    if campo is None:

        raise RuntimeError(
            "Input de arquivos HTML "
            "não encontrado."
        )

    print(
        f"📤 Enviando "
        f"{arquivo_html.name}"
    )

    campo.send_keys(
        str(
            arquivo_html.resolve()
        )
    )

    print(
        "✓ HTML enviado."
    )

    time.sleep(3)


# ============================================================
# LISTAR EXCELS
# ============================================================

def listar_excels():

    if not PASTA_EXCEL_NEXUS.exists():
        return set()

    return {
        arquivo.name
        for arquivo
        in PASTA_EXCEL_NEXUS.iterdir()
        if (
            arquivo.is_file()
            and
            arquivo.suffix.lower()
            in {
                ".xlsx",
                ".xls"
            }
            and
            not arquivo.name.startswith(
                "~$"
            )
        )
    }


# ============================================================
# LOCALIZAR BOTÃO EXCEL
# ============================================================

def localizar_botao_excel(
    driver
):

    elementos = driver.find_elements(
        By.XPATH,
        "//button | //a | //input"
    )

    for elemento in elementos:

        try:

            if not elemento.is_displayed():
                continue

            texto = normalizar_texto(
                elemento.text
            )

            valor = normalizar_texto(
                elemento.get_attribute(
                    "value"
                )
            )

            title = normalizar_texto(
                elemento.get_attribute(
                    "title"
                )
            )

            conteudo = (
                texto
                + " "
                + valor
                + " "
                + title
            )

            if "EXCEL" in conteudo:

                return elemento

        except Exception:

            continue

    return None


# ============================================================
# CLICAR EXCEL
# ============================================================

def clicar_excel(
    driver
):

    limite = (
        time.time()
        + TIMEOUT
    )

    botao = None

    while time.time() < limite:

        botao = localizar_botao_excel(
            driver
        )

        if botao:
            break

        time.sleep(0.5)

    if botao is None:

        raise RuntimeError(
            "Botão EXCEL não encontrado."
        )

    driver.execute_script(
        """
        arguments[0].scrollIntoView({
            block: 'center'
        });
        """,
        botao
    )

    time.sleep(0.5)

    try:

        botao.click()

    except Exception:

        driver.execute_script(
            "arguments[0].click();",
            botao
        )

    print(
        "✓ EXCEL clicado."
    )


# ============================================================
# ESPERAR DOWNLOAD EXCEL
# ============================================================

def esperar_download_excel(
    arquivos_antes
):

    limite = (
        time.time()
        + TIMEOUT_DOWNLOAD
    )

    while time.time() < limite:

        arquivos = [
            arquivo
            for arquivo
            in PASTA_EXCEL_NEXUS.iterdir()
            if arquivo.is_file()
        ]

        novos = [
            arquivo
            for arquivo in arquivos
            if (
                arquivo.name
                not in arquivos_antes
            )
            and
            arquivo.suffix.lower()
            in {
                ".xlsx",
                ".xls"
            }
            and
            not arquivo.name.startswith(
                "~$"
            )
        ]

        temporarios = [
            arquivo
            for arquivo in arquivos
            if (
                arquivo.name.endswith(
                    ".crdownload"
                )
                or
                arquivo.name.endswith(
                    ".tmp"
                )
            )
        ]

        if novos and not temporarios:

            arquivo = max(
                novos,
                key=lambda x:
                    x.stat().st_mtime
            )

            tamanho1 = (
                arquivo.stat().st_size
            )

            time.sleep(1)

            if arquivo.exists():

                tamanho2 = (
                    arquivo.stat().st_size
                )

                if tamanho1 == tamanho2:

                    return arquivo

        time.sleep(0.5)

    return None


# ============================================================
# RENOMEAR EXCEL
# ============================================================

def renomear_excel(
    arquivo,
    arquivo_html
):

    placa = arquivo_html.stem

    nome = (
        f"{nome_seguro(placa)}_"
        f"{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        f".xlsx"
    )

    destino = (
        PASTA_EXCEL_NEXUS
        / nome
    )

    if destino.exists():
        destino.unlink()

    shutil.move(
        str(arquivo),
        str(destino)
    )

    return destino


# ============================================================
# PROCESSAR HTML NO NEXUS
# ============================================================

def processar_html_nexus(
    driver,
    arquivo_html
):

    try:

        enviar_html(
            driver,
            arquivo_html
        )

        time.sleep(
            INTERVALO_ENTRE_ANALISES
        )

        arquivos_antes = listar_excels()

        clicar_excel(
            driver
        )

        print(
            "⬇ Aguardando Excel..."
        )

        arquivo = (
            esperar_download_excel(
                arquivos_antes
            )
        )

        if arquivo is None:

            raise TimeoutException(
                "Excel não baixado."
            )

        arquivo_final = (
            renomear_excel(
                arquivo,
                arquivo_html
            )
        )

        print(
            f"✓ Excel salvo: "
            f"{arquivo_final.name}"
        )

        return arquivo_final

    except Exception as erro:

        print(
            f"❌ Erro processando "
            f"{arquivo_html.name}: "
            f"{erro}"
        )

        return None


# ============================================================
# PROCESSAR TODOS OS HTMLS
# ============================================================

def processar_htmls_nexus(
    driver,
    arquivos_html
):

    arquivos_excel = []

    total = len(
        arquivos_html
    )

    for indice, arquivo_html in enumerate(
        arquivos_html,
        start=1
    ):

        print()

        print(
            "=" * 70
        )

        print(
            f"ANÁLISE {indice}/{total}"
        )

        print(
            arquivo_html.name
        )

        print(
            "=" * 70
        )

        arquivo_excel = (
            processar_html_nexus(
                driver,
                arquivo_html
            )
        )

        if arquivo_excel:

            arquivos_excel.append(
                arquivo_excel
            )

        time.sleep(
            INTERVALO_ENTRE_ANALISES
        )

    return arquivos_excel


# ============================================================
# COLUNA DESVIOS
# ============================================================

def encontrar_coluna_desvios(
    df
):

    for coluna in df.columns:

        if (
            normalizar_texto(
                coluna
            )
            == "DESVIOS"
        ):

            return coluna

    return None


# ============================================================
# FILTRO DOS DESVIOS
# ============================================================

def possui_desvio_permitido(
    valor
):

    texto = normalizar_texto(
        valor
    )

    if not texto:
        return False

    desvios = [
        "MOTOR LIGA",
        "DIVERGENCIA ENDERECO",
        "MOTORISTA DIVERGENTE",
    ]

    for desvio in desvios:

        if desvio in texto:
            return True

    return False


def filtrar_desvios(
    df
):

    coluna = encontrar_coluna_desvios(
        df
    )

    if coluna is None:
        return pd.DataFrame()

    mascara = (
        df[coluna]
        .fillna("")
        .map(
            possui_desvio_permitido
        )
    )

    return df.loc[
        mascara
    ].copy()


# ============================================================
# LER EXCEL
# ============================================================

def ler_excel_desvios(
    arquivo
):

    registros = []

    try:

        xls = pd.ExcelFile(
            arquivo
        )

    except Exception as erro:

        print(
            f"⚠ Erro em {arquivo.name}: "
            f"{erro}"
        )

        return pd.DataFrame()

    for aba in xls.sheet_names:

        try:

            df = pd.read_excel(
                arquivo,
                sheet_name=aba
            )

            if df.empty:
                continue

            df_filtrado = (
                filtrar_desvios(
                    df
                )
            )

            if df_filtrado.empty:
                continue

            df_filtrado.insert(
                0,
                "ARQUIVO_ORIGEM",
                arquivo.name
            )

            df_filtrado.insert(
                1,
                "ABA_ORIGEM",
                aba
            )

            registros.append(
                df_filtrado
            )

        except Exception as erro:

            print(
                f"⚠ Erro na aba "
                f"{aba}: {erro}"
            )

    if not registros:
        return pd.DataFrame()

    return pd.concat(
        registros,
        ignore_index=True,
        sort=False
    )


# ============================================================
# FORMATAR EXCEL
# ============================================================

def formatar_excel(
    arquivo,
    aba
):

    wb = load_workbook(
        arquivo
    )

    ws = wb[
        aba
    ]

    for cell in ws[1]:

        cell.font = Font(
            bold=True
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    ws.freeze_panes = "A2"

    if ws.max_row >= 1:

        ws.auto_filter.ref = (
            ws.dimensions
        )

    for coluna in ws.columns:

        maior = 0

        for cell in coluna:

            valor = (
                ""
                if cell.value is None
                else str(cell.value)
            )

            maior = max(
                maior,
                len(valor)
            )

        letra = get_column_letter(
            coluna[0].column
        )

        ws.column_dimensions[
            letra
        ].width = min(
            max(
                maior + 2,
                10
            ),
            60
        )

    wb.save(
        arquivo
    )


# ============================================================
# UNIFICAR DESVIOS
# ============================================================

def gerar_unificado_desvios():

    print(
        "\n" + "=" * 70
    )

    print(
        "GERANDO UNIFICADO DE DESVIOS"
    )

    print(
        "=" * 70
    )

    todos = []

    if ARQUIVO_AUDITORIA.exists():

        try:

            xls = pd.ExcelFile(
                ARQUIVO_AUDITORIA
            )

            for aba in [
                "Desvios",
                "Abastecimento_x_Telemetria",
                "Paradas_Motor_Ligado",
            ]:

                if aba not in xls.sheet_names:
                    continue

                df = pd.read_excel(
                    ARQUIVO_AUDITORIA,
                    sheet_name=aba
                )

                if not df.empty:

                    df[
                        "ABA_ORIGEM"
                    ] = aba

                    todos.append(
                        df
                    )

        except Exception as erro:

            print(
                f"⚠ Erro lendo auditoria: "
                f"{erro}"
            )

    if todos:

        final = (
            pd.concat(
                todos,
                ignore_index=True,
                sort=False
            )
            .drop_duplicates()
            .reset_index(drop=True)
        )

    else:

        final = pd.DataFrame(
            columns=[
                "PLACA",
                "DATA_HORA",
                "TIPO_DESVIO",
                "CRITICIDADE",
                "DETALHE",
                "ABA_ORIGEM",
            ]
        )

    with pd.ExcelWriter(
        ARQUIVO_UNIFICADO,
        engine="openpyxl"
    ) as writer:

        final.to_excel(
            writer,
            sheet_name="Desvios",
            index=False
        )

    formatar_excel(
        ARQUIVO_UNIFICADO,
        "Desvios"
    )

    print(
        f"✓ UNIFICADO GERADO: "
        f"{ARQUIVO_UNIFICADO.resolve()}"
    )

    print(
        f"✓ Registros: "
        f"{len(final)}"
    )

    return ARQUIVO_UNIFICADO


# ============================================================
# MAIN
# ============================================================

def main():

    print()

    print(
        "=" * 70
    )

    print(
        "AUTOVISION → AUDITORIA DE TELEMETRIA"
    )

    print(
        "=" * 70
    )

    # ========================================================
    # VALIDAR PLACAS
    # ========================================================

    global PLACAS

    PLACAS = carregar_placas_telemetria_ociosidade()


    
    # ========================================================
    # PERÍODO
    # ========================================================

    inicio, fim = obter_periodo()

    driver = criar_driver()

    try:

        # ====================================================
        # AUTOVISION
        # ====================================================

        fazer_login(
            driver
        )

        arquivos_html = (
            processar_placas(
                driver,
                inicio,
                fim
            )
        )

        print()

        print(
            "=" * 70
        )

        print(
            "RESULTADO AUTOVISION"
        )

        print(
            "=" * 70
        )

        print(
            f"Placas configuradas: "
            f"{len(PLACAS)}"
        )

        print(
            f"HTMLs baixados: "
            f"{len(arquivos_html)}"
        )

        if not arquivos_html:

            arquivos_html = (
                listar_htmls_periodo(
                    inicio,
                    fim
                )
            )

        if not arquivos_html:

            print(
                "❌ Nenhum HTML foi baixado "
                "ou encontrado na pasta."
            )

            return

        # ====================================================
        # AUDITORIA LOCAL
        # ====================================================

        arquivos_para_analise = list(
            arquivos_html
        )

        existentes = (
            listar_htmls_periodo(
                inicio,
                fim
            )
        )

        nomes = {
            a.name
            for a in arquivos_para_analise
        }

        for arquivo in existentes:

            if arquivo.name not in nomes:

                arquivos_para_analise.append(
                    arquivo
                )

        executar_auditoria(
            arquivos_para_analise,
            inicio,
            fim
        )

        arquivo_unificado = gerar_unificado_desvios()

        # ====================================================
        # ENVIO PARA O GOOGLE SHEETS
        # ====================================================

        try:

            df_resultados = pd.read_excel(
                arquivo_unificado,
                sheet_name="Desvios"
            )

            enviar_resultados_google_sheets(
                df_resultados
            )

        except Exception as erro_envio:

            print(
                f"⚠ Erro ao enviar resultados para o "
                f"Google Sheets: {erro_envio}"
            )

        # ====================================================
        # FINAL
        # ====================================================

        print()

        print(
            "=" * 70
        )

        print(
            "✅ PROCESSO FINALIZADO"
        )

        print(
            "=" * 70
        )

        print(
            f"HTMLs: "
            f"{PASTA_HTML.resolve()}"
        )

        print(
            f"Auditoria: "
            f"{ARQUIVO_AUDITORIA.resolve()}"
        )

        print(
            f"Unificado: "
            f"{ARQUIVO_UNIFICADO.resolve()}"
        )

    except KeyboardInterrupt:

        print(
            "\nProcesso interrompido."
        )

    except Exception as erro:

        print()

        print(
            "=" * 70
        )

        print(
            "❌ ERRO GERAL"
        )

        print(
            "=" * 70
        )

        print(
            erro
        )

        import traceback

        traceback.print_exc()

    finally:

        try:

            driver.quit()

        except Exception:

            pass




# ============================================================
# ENVIO DOS RESULTADOS PARA O GOOGLE SHEETS
# ============================================================

def enviar_resultados_google_sheets(resultados):
    """
    Envia os resultados processados para o Google Apps Script,
    que grava os dados na aba RESULTADOS.
    """

    print("\n" + "=" * 60)
    print("ENVIANDO RESULTADOS PARA O GOOGLE SHEETS")
    print("=" * 60)

    try:
        # URL do seu Web App do Google Apps Script
        url = GOOGLE_SCRIPT_URL

        # Converte DataFrame para lista de dicionários
        if isinstance(resultados, pd.DataFrame):
            dados = resultados.fillna("").to_dict(orient="records")
        else:
            dados = resultados

        payload = {
            "acao": "resultados",
            "resultados": dados
        }

        response = requests.post(
            url,
            json=payload,
            timeout=60
        )

        response.raise_for_status()

        print("\nRESULTADOS ENVIADOS COM SUCESSO!")
        print(response.text)

        return True

    except requests.exceptions.RequestException as e:
        print("\nERRO DE COMUNICAÇÃO:")
        print(e)

        return False

    except Exception as e:
        print("\nERRO AO ENVIAR RESULTADOS:")
        print(e)

        return False

















# ============================================================
# EXECUTAR
# ============================================================

if __name__ == "__main__":

    main()
